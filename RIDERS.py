#!/usr/bin/env python3
"""
Freeride Fanatics — Instagram Rider Card Generator
====================================================
Génère les visuels Instagram pour chaque rider à partir du tracker Excel.
 
STRUCTURE DOSSIERS REQUISE :
  generate_cards.py         ← ce script
  assets/
    background.png          ← fond exporté depuis Figma (1080×1350)
  fonts/
    BebasNeue-Regular.ttf   ← https://fonts.google.com/specimen/Bebas+Neue
    AntonSC-Regular.ttf     ← https://fonts.google.com/specimen/Anton+SC
  photos/
    goldstone_jackson.png   ← photos riders (format: nom_prenom.png en minuscules)
    vergier_loris.png
    ...
  output/                   ← créé automatiquement, cartes générées ici
  UCI_DH_2026_Tracker_v3.xlsx
 
INSTALLATION :
  pip install pillow openpyxl instagrapi
 
UTILISATION :
  python3 generate_cards.py              # génère toutes les cartes
  python3 generate_cards.py --post       # génère + publie sur Instagram
  python3 generate_cards.py --rider GOLDSTONE  # un seul rider
"""
 
import os
import sys
import argparse
from pathlib import Path
from datetime import date
from PIL import Image, ImageDraw, ImageFont
import openpyxl
 
# ── CONFIGURATION ─────────────────────────────────────────────────────────────
 
BASE_DIR    = Path(__file__).parent
ASSETS_DIR  = BASE_DIR / "assets"
FONTS_DIR   = BASE_DIR / "fonts"
PHOTOS_DIR  = BASE_DIR / "photos"
OUTPUT_DIR  = BASE_DIR / "output"
EXCEL_FILE  = BASE_DIR / "UCI_DH_2026_Tracker_v3.xlsx"
 
BACKGROUND  = ASSETS_DIR / "background.png"
 
# Polices
FONT_LABEL  = FONTS_DIR / "AntonSC-Regular.ttf"    # titres jaunes
FONT_VALUE  = FONTS_DIR / "BebasNeue-Regular.ttf"  # valeurs blanches
 
# Couleurs (reprises du template Figma)
C_LIME      = "#C8D400"   # labels jaunes
C_WHITE     = "#FFFFFF"   # valeurs blanches
C_PANEL     = "#4A4A4A"   # panneau gris droit (fallback si pas de background.png)
C_BORDER    = "#B8CC00"   # bordure verte
 
# Dimensions carte
W, H        = 1080, 1350
 
# Zone photo (gauche) — zone blanche
PHOTO_X     = 0
PHOTO_Y     = 0
PHOTO_W     = 430      # largeur zone photo
PHOTO_H     = H
 
# Zone texte (droite) — panneau gris
TEXT_X      = 490      # marge gauche texte dans le panneau
TEXT_TOP    = 90       # marge haute première section
LINE_GAP    = 180     # espacement vertical entre chaque section
 
# Tailles de police
SZ_LABEL    = 38       # Anton SC (labels jaunes)
SZ_VALUE    = 58       # Bebas Neue (valeurs blanches, noms)
SZ_VALUE_SM = 44       # Bebas Neue valeurs secondaires
SZ_PALMARES = 36       # Bebas Neue palmarès (multi-lignes)
 
# ── CHARGEMENT POLICES ────────────────────────────────────────────────────────
 
def load_fonts():
    """Charge les polices. Utilise Poppins Bold en fallback si absentes."""
    fallback = "/usr/share/fonts/truetype/google-fonts/Poppins-Bold.ttf"
 
    def get_font(path, size):
        try:
            return ImageFont.truetype(str(path), size)
        except:
            print(f"⚠️  Police non trouvée : {path.name} → fallback Poppins")
            try:
                return ImageFont.truetype(fallback, size)
            except:
                return ImageFont.load_default()
 
    return {
        "label":      get_font(FONT_LABEL, SZ_LABEL),
        "value":      get_font(FONT_VALUE, SZ_VALUE),
        "value_sm":   get_font(FONT_VALUE, SZ_VALUE_SM),
        "palmares":   get_font(FONT_VALUE, SZ_PALMARES),
    }
 
# ── CHARGEMENT DONNÉES EXCEL ──────────────────────────────────────────────────
 
def load_profiles():
    """Charge la feuille 👤 Profils du tracker Excel."""
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["👤 Profils"]
    profiles = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] in ("F", "M") and row[2]:  # genre + nom présent
            profiles.append({
                "genre":       row[0],
                "prenom":      row[1] or "",
                "nom":         row[2] or "",
                "flag":        row[3] or "",
                "nationalite": row[4] or "À confirmer",
                "ville":       row[5] or "À confirmer",
                "birthday":    row[6] or "À confirmer",
                "age":         str(row[7]) if row[7] else "?",
                "palmares":    row[8] or "À confirmer",
                "category":    "Women Elite" if row[0] == "F" else "Men Elite",
            })
 
    # Ajoute l'équipe depuis la feuille 🏆 Riders
    ws2 = wb["🏆 Riders"]
    teams = {}
    for row in ws2.iter_rows(min_row=3, values_only=True):
        if row[0] in ("F", "M") and row[2]:
            teams[row[2]] = row[5] or "À confirmer"  # nom → equipe
 
    for p in profiles:
        p["team"] = teams.get(p["nom"], "À confirmer")
 
    return profiles
 
# ── GÉNÉRATION CARTE ──────────────────────────────────────────────────────────
 
def find_photo(prenom, nom):
    """Cherche la photo du rider dans le dossier photos/."""
    # Formats tentés : nom_prenom.png / prenom_nom.png / nom.png
    stems = [
        f"{nom.lower().replace(' ', '_')}_{prenom.lower().replace('-', '_')}",
        f"{prenom.lower().replace('-', '_')}_{nom.lower().replace(' ', '_')}",
        f"{nom.lower().replace(' ', '_')}",
    ]
    for stem in stems:
        for ext in [".png", ".jpg", ".jpeg"]:
            p = PHOTOS_DIR / (stem + ext)
            if p.exists():
                return p
    return None
 
 
def draw_text_wrapped(draw, text, font, color, x, y, max_width):
    """Dessine un texte avec retour à la ligne automatique."""
    words = text.split()
    lines = []
    current = ""
    for word in words:
        test = (current + " " + word).strip()
        bbox = draw.textbbox((0, 0), test, font=font)
        if bbox[2] - bbox[0] <= max_width:
            current = test
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
 
    line_h = font.size + 6
    for line in lines:
        draw.text((x, y), line, font=font, fill=color)
        y += line_h
    return y  # retourne y final
 
 
def generate_card(profile, fonts, bg_image):
    """Génère une carte Instagram pour un rider."""
    # Base : fond du template
    card = bg_image.copy().convert("RGBA")
    draw = ImageDraw.Draw(card)
 
    # ── Photo rider ──────────────────────────────────────────────────────────
    photo_path = find_photo(profile["prenom"], profile["nom"])
    if photo_path:
        try:
            photo = Image.open(photo_path).convert("RGBA")
            # Redimensionne pour remplir la zone blanche (crop centré)
            ratio = max(PHOTO_W / photo.width, PHOTO_H / photo.height)
            new_w = int(photo.width * ratio)
            new_h = int(photo.height * ratio)
            photo = photo.resize((new_w, new_h), Image.LANCZOS)
            # Centre le crop
            crop_x = (new_w - PHOTO_W) // 2
            crop_y = (new_h - PHOTO_H) // 2
            photo = photo.crop((crop_x, crop_y, crop_x + PHOTO_W, crop_y + PHOTO_H))
            card.paste(photo, (PHOTO_X, PHOTO_Y), photo)
        except Exception as e:
            print(f"  ⚠️  Photo erreur ({photo_path.name}): {e}")
    else:
        print(f"  ℹ️  Pas de photo trouvée pour {profile['prenom']} {profile['nom']}")
 
    # ── Textes ───────────────────────────────────────────────────────────────
    MAX_TXT_W = W - TEXT_X - 20   # largeur max texte dans le panneau
 
    sections = [
        ("Name",        f"{profile['prenom']} {profile['nom']}",   "value"),
        ("Nationality", profile["nationalite"],                      "value_sm"),
        ("Birthday",    f"{profile['birthday']}  ·  {profile['age']} ans", "value_sm"),
        ("Homestay",    profile["ville"],                            "value_sm"),
        ("Category",    profile["category"],                         "value_sm"),
        ("Team",        profile["team"],                             "value_sm"),
        ("Palmares",    profile["palmares"],                         "palmares"),
    ]
 
    y = TEXT_TOP
    for label, value, font_key in sections:
        # Label (Anton SC, lime)
        draw.text((TEXT_X, y), label.upper(), font=fonts["label"], fill=C_LIME)
        y += fonts["label"].size + 6
 
        # Valeur (Bebas Neue, blanc)
        y = draw_text_wrapped(draw, value.upper(), fonts[font_key], C_WHITE,
                              TEXT_X, y, MAX_TXT_W)
        y += 22   # espace après valeur
 
    # ── Drapeau ──────────────────────────────────────────────────────────────
    # (emoji drapeau en bas du panneau — optionnel, désactivez si l'emoji ne s'affiche pas)
    # draw.text((TEXT_X, H - 80), profile["flag"], font=fonts["value"], fill=C_WHITE)
 
    return card.convert("RGB")
 
 
def make_background_fallback():
    """Crée un fond de remplacement si background.png est absent."""
    img = Image.new("RGB", (W, H), "#FFFFFF")
    draw = ImageDraw.Draw(img)
 
    # Panneau gris droit avec bord irrégulier (simulation brush stroke)
    import math, random
    random.seed(42)
    panel_x = 415
 
    # Dessine le panneau gris
    panel_pts = [(panel_x + random.randint(-18, 18), y) for y in range(0, H + 10, 8)]
    poly = [(W, 0)] + panel_pts[::-1] + [(W, H)]
    draw.polygon([(W, 0)] + [(panel_x, 0)] + panel_pts + [(panel_x, H), (W, H)],
                 fill=C_PANEL)
 
    # Bordure lime
    border = 8
    draw.rectangle([border//2, border//2, W - border//2, H - border//2],
                   outline=C_BORDER, width=border)
 
    return img
 
 
# ── PUBLICATION INSTAGRAM ─────────────────────────────────────────────────────
 
def post_to_instagram(image_path, caption, username, password):
    """Publie une image sur Instagram via Instagrapi."""
    try:
        from instagrapi import Client
    except ImportError:
        print("⚠️  instagrapi non installé — lancez : pip install instagrapi")
        return False
 
    cl = Client()
    try:
        cl.login(username, password)
        cl.photo_upload(str(image_path), caption)
        print(f"  ✅ Publié : {image_path.name}")
        return True
    except Exception as e:
        print(f"  ❌ Erreur publication : {e}")
        return False
 
 
def build_caption(profile):
    """Génère la caption Instagram pour un rider."""
    flag = profile["flag"]
    name = f"{profile['prenom']} {profile['nom']}"
    team = profile["team"]
    palmares = profile["palmares"]
    category = profile["category"]
 
    caption = f"{flag} {name}\n"
    caption += f"🏔️ {team}\n"
    caption += f"🏆 {palmares}\n\n"
    caption += f"#UCI #DHWorldCup #DH2026 #{profile['nom'].replace(' ', '')} "
    caption += f"#DownhillMTB #MTB #freeride #mountainbike #{category.replace(' ', '')}"
    return caption
 
 
# ── MAIN ──────────────────────────────────────────────────────────────────────
 
def main():
    parser = argparse.ArgumentParser(description="Génère les cartes Instagram DH 2026")
    parser.add_argument("--rider",  help="Génère uniquement ce rider (nom en majuscules, ex: GOLDSTONE)")
    parser.add_argument("--post",   action="store_true", help="Publie sur Instagram après génération")
    parser.add_argument("--user",   help="Nom d'utilisateur Instagram (si --post)")
    parser.add_argument("--pass",   dest="pwd", help="Mot de passe Instagram (si --post)")
    parser.add_argument("--women",  action="store_true", help="Women Elite uniquement")
    parser.add_argument("--men",    action="store_true", help="Men Elite uniquement")
    args = parser.parse_args()
 
    OUTPUT_DIR.mkdir(exist_ok=True)
 
    # Charge le fond
    if BACKGROUND.exists():
        bg = Image.open(BACKGROUND).convert("RGBA").resize((W, H), Image.LANCZOS)
        print(f"✅ Fond chargé : {BACKGROUND.name}")
    else:
        print("⚠️  background.png absent → fond généré automatiquement")
        bg = make_background_fallback().convert("RGBA")
 
    # Charge les polices
    fonts = load_fonts()
 
    # Charge les données
    profiles = load_profiles()
    print(f"✅ {len(profiles)} riders chargés depuis le tracker")
 
    # Filtre
    if args.rider:
        profiles = [p for p in profiles if args.rider.upper() in p["nom"].upper()]
    if args.women:
        profiles = [p for p in profiles if p["genre"] == "F"]
    if args.men:
        profiles = [p for p in profiles if p["genre"] == "M"]
 
    print(f"→ Génération de {len(profiles)} carte(s)...\n")
 
    for p in profiles:
        name_slug = f"{p['nom'].lower().replace(' ', '_')}_{p['prenom'].lower()}"
        out_path = OUTPUT_DIR / f"{name_slug}.jpg"
 
        print(f"  🎨 {p['prenom']} {p['nom']} ({p['category']})")
        card = generate_card(p, fonts, bg)
        card.save(out_path, "JPEG", quality=95)
        print(f"     → {out_path.name}")
 
        if args.post and args.user and args.pwd:
            caption = build_caption(p)
            post_to_instagram(out_path, caption, args.user, args.pwd)
 
    print(f"\n✅ Terminé — cartes dans : {OUTPUT_DIR}/")
 
 
if __name__ == "__main__":
    main()
 