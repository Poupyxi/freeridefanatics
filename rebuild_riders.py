from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()

C_BG      = '1A1A2E'
C_LIME    = 'C8D400'
C_PURPLE  = '4A1942'
C_DARK    = '111111'
C_MEN_A   = 'F5F5F5'
C_MEN_B   = 'FFFFFF'
C_WOM_A   = 'FFF0F5'
C_WOM_B   = 'FFF8FC'

def fill(c): return PatternFill('solid', fgColor=c)
def font(c='111111', bold=False, sz=10): return Font(name='Arial', color=c, bold=bold, size=sz)
def align(h='left', v='center', wrap=True): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style(cell, bg, fg=C_DARK, bold=False, sz=10, h='left'):
    cell.fill = fill(bg)
    cell.font = font(fg, bold, sz)
    cell.alignment = align(h)

def header_row(ws, cols, bg, fg, bold=True, sz=10):
    for col, val in enumerate(cols, 1):
        c = ws.cell(1, col, val)
        style(c, bg, fg, bold, sz, 'center')

def title_row(ws, val, ncols, bg, fg, bold=True, sz=13):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, val)
    style(c, bg, fg, bold, sz, 'center')

def section_row(ws, row, val, ncols, bg=C_PURPLE, fg=C_LIME):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, val)
    style(c, bg, fg, bold=True, sz=11, h='center')

def data_row(ws, row, values, bg):
    for col, val in enumerate(values, 1):
        c = ws.cell(row, col, val)
        style(c, bg)

# ═══════════════════════════════════════════════════════════════
# SHEET 1 — Profils
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = '👤 Profils'
ws1.row_dimensions[1].height = 30

NCOLS_P = 11

# Row 1: title
ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS_P)
c = ws1.cell(1, 1, 'FREERIDE FANATICS — RIDER PROFILES 2026')
style(c, C_BG, C_LIME, bold=True, sz=13, h='center')

# Row 2: headers
headers_p = ['Genre','Prénom','Nom','Flag','Nationalité','Ville',
             'Date de naissance','Âge','Palmarès','Team','Instagram']
for col, h in enumerate(headers_p, 1):
    c = ws1.cell(2, col, h)
    style(c, C_BG, C_LIME, bold=True, sz=10, h='center')

# Column widths
widths_p = [7, 14, 16, 6, 16, 22, 16, 6, 50, 30, 18]
for i, w in enumerate(widths_p, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

MEN = [
    ('M','Martin','MAES','🇧🇪','Belgique','Liège','27 Jan 1997',29,
     'WC Win 2018 · World Champs 🥈 2018 · Euro 4X Champ 2010','Orbea FMD Racing','@martin_maes5'),
    ('M','Max','ALRAN','🇫🇷','France','France','23 Mar 2007',19,
     'Junior World Champion 2025 · French & Euro U17 DH Champion 2023','Commencal/Muc-Off by Riding Addiction','@maxalran'),
    ('M','Andreas','KOLB','🇦🇹','Autriche','Schladming','23 Mar 1996',30,
     'WC Win Leogang 2023 · World Champs 🥈 Fort William 2023 · European Champion 2024','Santa Cruz Syndicate','@andikolb'),
    ('M','Evan','MEDCALF','🇺🇸','États-Unis','Albuquerque, NM','01 Jun 2005',21,
     'Vice US Champion 2025 · 1er Top 15 WC','Commencal Schwalbe by Les Orres','@evan.medcalf'),
    ('M','Aaron','GWIN','🇺🇸','États-Unis','Morongo Valley, CA','24 Déc 1987',38,
     '5x WC Overall Champion · 5 WC Wins en 2011 · Légende du DH','Frameworks Racing / TRP','@aarongwin'),
    ('M','Roger','VIEIRA','🇧🇷','Brésil','Brésil / UK','03 Nov 1994',31,
     'Cerro Abajo Genoa 2025 · Champion Brésil 2024 · Elite British Champion 2021','Pivot Factory Racing','@rogervieira8'),
    ('M','Richard','RUDE JR','🇺🇸','États-Unis','Redding, Connecticut','06 Fév 1995',31,
     'Junior DH World Champ 2013 · 3x EWS Champion (2015, 2016, 2018)','Yeti / Fox Factory Race Team','@richie_rude1'),
    ('M','George','MADLEY','🇬🇧','Grande-Bretagne','Grande-Bretagne','28 Fév 2007',19,
     'Junior British National Champion · World Champs 🥉 2025','Continental Atherton','@george.madley.mtb'),
    ('M','Loïc','BRUNI','🇫🇷','France',"Côte d'Azur",'13 Mai 1994',32,
     '5x World Champion (2015,2019,2020,2021,2022) · 4x WC Overall','Specialized Gravity','@loicbruni29'),
    ('M','Luke','WALKER','🇦🇺','Australie','Australie','À confirmer',None,
     'Privateer AUS · résultats nationaux','Privateer','@lukewalkerracing'),
    ('M','Bernard','KERR','🇬🇧','Grande-Bretagne','Guildford, Angleterre','16 Fév 1991',35,
     '3x Red Bull Hardline Winner · Multiple WC Podiums · Founder Pivot Factory Racing','Pivot Factory Racing','@bernard_kerr'),
]

WOMEN = [
    ('F','Valentina','ROA SANCHEZ','🇨🇴','Colombie','Ibagué','04 Mar 2005',21,
     'Junior World Champion 2023 · 1ère Latino-Américaine WC DH · 2x Continental Champion','MS Racing','@valenrs5'),
    ('F','Lisa','BOULADOU','🇫🇷','France','France','04 Jan 2005',21,
     '2x Championne de France (2024) · Junior WC Win Leogang 2024','Santa Cruz / Burgtec by Goodman','@lisa_bouladou'),
    ('F','Tahnée','SEAGRAVE','🇬🇧','Grande-Bretagne',"Saint-Jean-d'Aulps / Wales",'15 Juin 1995',30,
     'Multiple WC Wins · Junior World Champ · 3 WC Wins 2024-2025','Orbea FMD Racing','@tahneeseagrave'),
    ('F','Gracey','HEMSTREET','🇨🇦','Canada','Sechelt, BC','18 Nov 2004',21,
     '1ère femme Red Bull Hardline (2x) · 1ère Canadienne WC winner · WC Loudenvielle, Leogang, Les Gets 2025','Norco Race Division','@graceyhemstreet'),
    ('F','Harriet','HARNDEN','🇬🇧','Grande-Bretagne','West Malvern','09 Mar 2001',25,
     '77 victoires · EWS Overall 2024 · UK DH National Champ 2023 · British CX Champ 2020 & 2022','AON Racing','@hatt1e_harnden'),
    ('F','Jess','BLEWITT','🇳🇿','Nouvelle-Zélande','Mt Maunganui','À confirmer',21,
     '2x NZ National Champion · Crankworx Rotorua 2023 · 1ère femme Red Bull Hardline','Scott DH Factory Racing','@jessblewitt_'),
    ('F','Frida Helena','RØNNING','🇳🇴','Norvège','Lillehammer','1995',31,
     '3x Championne de Norvège','Merida NTG','@frida9'),
    ('F','Laïs','BONNAURE','🇫🇷','France','Luberon, France','17 Mai 2006',20,
     "Championne de France 2022 & 2024 · Championne d'Europe U23 2023",'VTT Luberon','@lais_bonnaure'),
]

row = 3
section_row(ws1, row, '— MEN ELITE —', NCOLS_P, C_BG, C_LIME)
ws1.row_dimensions[row].height = 22
row += 1
for i, d in enumerate(MEN):
    bg = C_MEN_A if i % 2 == 0 else C_MEN_B
    data_row(ws1, row, d, bg)
    ws1.row_dimensions[row].height = 40
    row += 1

section_row(ws1, row, '— WOMEN ELITE —', NCOLS_P, C_PURPLE, C_LIME)
ws1.row_dimensions[row].height = 22
row += 1
for i, d in enumerate(WOMEN):
    bg = C_WOM_A if i % 2 == 0 else C_WOM_B
    data_row(ws1, row, d, bg)
    ws1.row_dimensions[row].height = 40
    row += 1

# ═══════════════════════════════════════════════════════════════
# SHEET 2 — Riders & Equipment
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('🏆 Riders & Equipment')
NCOLS_E = 14

ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS_E)
c = ws2.cell(1, 1, 'FREERIDE FANATICS — RIDERS & EQUIPMENT 2026')
style(c, C_BG, C_LIME, bold=True, sz=13, h='center')

headers_e = ['Genre','Prénom','Nom','Flag','Team','Vélo','Fourche','Amortisseur',
             'Freins','Roues','Pneus','Casque','Masque','Instagram']
for col, h in enumerate(headers_e, 1):
    c = ws2.cell(2, col, h)
    style(c, C_BG, C_LIME, bold=True, sz=10, h='center')

widths_e = [7,14,16,6,30,18,12,14,10,14,12,12,12,18]
for i, w in enumerate(widths_e, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

MEN_E = [
    ('M','Martin','MAES','🇧🇪','Orbea FMD Racing','Orbea','Fox 40','Fox DHX2','SRAM',None,'Maxxis',None,None,'@martin_maes5'),
    ('M','Max','ALRAN','🇫🇷','Commencal/Muc-Off by Riding Addiction','Commencal','Fox 40','Fox DHX2','SRAM',None,'Schwalbe',None,None,'@maxalran'),
    ('M','Andreas','KOLB','🇦🇹','Santa Cruz Syndicate','Santa Cruz V10','Fox 40','Fox DHX2','SRAM','DT Swiss','Maxxis','Poc','Oakley','@andikolb'),
    ('M','Evan','MEDCALF','🇺🇸','Commencal Schwalbe by Les Orres','Commencal','Fox 40','Fox DHX2','SRAM',None,'Schwalbe',None,None,'@evan.medcalf'),
    ('M','Aaron','GWIN','🇺🇸','Frameworks Racing / TRP','Frameworks','Fox 40','Fox DHX2','TRP','ENVE','Continental','Bell','Oakley','@aarongwin'),
    ('M','Roger','VIEIRA','🇧🇷','Pivot Factory Racing','Pivot Phoenix','Fox 40','Fox DHX2','SRAM','Reserve','Maxxis','Troy Lee','Oakley','@rogervieira8'),
    ('M','Richard','RUDE JR','🇺🇸','Yeti / Fox Factory Race Team','Yeti','Fox 38','Fox DHX2','SRAM','Industry Nine','Maxxis','Kali','Oakley','@richie_rude1'),
    ('M','George','MADLEY','🇬🇧','Continental Atherton','Atherton','Fox 40','Fox DHX2','SRAM',None,'Maxxis',None,None,'@george.madley.mtb'),
    ('M','Loïc','BRUNI','🇫🇷','Specialized Gravity','Specialized Demo','Fox 40','Fox DHX2','SRAM','Roval','Maxxis','Poc','Oakley','@loicbruni29'),
    ('M','Luke','WALKER','🇦🇺','Privateer',None,None,None,None,None,'Maxxis',None,None,'@lukewalkerracing'),
    ('M','Bernard','KERR','🇬🇧','Pivot Factory Racing','Pivot Phoenix','Fox 40','Fox DHX2','SRAM',None,'Maxxis','Troy Lee','Oakley','@bernard_kerr'),
]

WOMEN_E = [
    ('F','Valentina','ROA SANCHEZ','🇨🇴','MS Racing','Zerode G3',None,None,None,None,None,None,None,'@valenrs5'),
    ('F','Lisa','BOULADOU','🇫🇷','Santa Cruz / Burgtec by Goodman','Santa Cruz V10','Fox 40','Fox DHX2','SRAM','Burgtec','Maxxis',None,None,'@lisa_bouladou'),
    ('F','Tahnée','SEAGRAVE','🇬🇧','Orbea FMD Racing','Orbea Rallon','Fox 40','Fox DHX2','SRAM',None,'Maxxis','Fox Racing','Oakley','@tahneeseagrave'),
    ('F','Gracey','HEMSTREET','🇨🇦','Norco Race Division','Norco','Fox 40','Fox DHX2','SRAM',None,'Maxxis',None,None,'@graceyhemstreet'),
    ('F','Harriet','HARNDEN','🇬🇧','AON Racing',None,'Fox 40','Fox DHX2','SRAM',None,'Maxxis',None,None,'@hatt1e_harnden'),
    ('F','Jess','BLEWITT','🇳🇿','Scott DH Factory Racing','Scott Gambler','Fox 40','Fox DHX2','SRAM',None,'Maxxis',None,None,'@jessblewitt_'),
    ('F','Frida Helena','RØNNING','🇳🇴','Merida NTG','Merida One-Sixty','Fox 40','Fox DHX2','SRAM',None,'Maxxis',None,None,'@frida9'),
    ('F','Laïs','BONNAURE','🇫🇷','VTT Luberon',None,'Fox 36','Fox DHX2','SRAM',None,'Maxxis',None,None,'@lais_bonnaure'),
]

row = 3
section_row(ws2, row, '— MEN ELITE —', NCOLS_E, C_BG, C_LIME)
ws2.row_dimensions[row].height = 22
row += 1
for i, d in enumerate(MEN_E):
    bg = C_MEN_A if i % 2 == 0 else C_MEN_B
    data_row(ws2, row, d, bg)
    ws2.row_dimensions[row].height = 22
    row += 1

section_row(ws2, row, '— WOMEN ELITE —', NCOLS_E, C_PURPLE, C_LIME)
ws2.row_dimensions[row].height = 22
row += 1
for i, d in enumerate(WOMEN_E):
    bg = C_WOM_A if i % 2 == 0 else C_WOM_B
    data_row(ws2, row, d, bg)
    ws2.row_dimensions[row].height = 22
    row += 1

wb.save('/sessions/dreamy-vibrant-volta/mnt/freeride/Riders_Selection_2026.xlsx')
print("✅ Riders_Selection_2026.xlsx créé avec 19 riders (11M + 8F)")
