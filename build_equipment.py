from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = '🔧 Équipements'

C_BG   = '1A1A2E'
C_LIME = 'C8D400'
C_PURPLE = '4A1942'
C_MEN_A  = 'F5F5F5'
C_MEN_B  = 'FFFFFF'
C_WOM_A  = 'FFF0F5'
C_WOM_B  = 'FFF8FC'
C_DARK   = '111111'

def fill(c): return PatternFill('solid', fgColor=c)
def font(c='111111', bold=False, sz=10): return Font(name='Arial', color=c, bold=bold, size=sz)
def align(h='center', v='center', wrap=True): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style(c, bg, fg=C_DARK, bold=False, sz=10, h='center'):
    c.fill = fill(bg)
    c.font = font(fg, bold, sz)
    c.alignment = align(h)

HEADERS = [
    'Genre', 'Prénom', 'Nom', 'Flag', 'Team',
    'Vélo / Cadre', 'Fourche', 'Amortisseur', 'Freins',
    'Transmission', 'Roues', 'Pneus', 'Casque', 'Masque',
    'Cockpit / Guidon', 'Notes', 'Source'
]
NCOLS = len(HEADERS)

# Row 1: title
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
c = ws.cell(1, 1, 'FREERIDE FANATICS — ÉQUIPEMENTS RIDERS 2025/2026')
style(c, C_BG, C_LIME, bold=True, sz=13)
ws.row_dimensions[1].height = 28

# Row 2: headers
for col, h in enumerate(HEADERS, 1):
    c = ws.cell(2, col, h)
    style(c, C_BG, C_LIME, bold=True, sz=9)
ws.row_dimensions[2].height = 22

# Column widths
widths = [7,14,16,5,28,20,14,16,14,14,18,22,12,12,22,30,30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

def section(ws, row, label, ncols, bg=C_BG, fg=C_LIME):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row, 1, label)
    style(c, bg, fg, bold=True, sz=11)
    ws.row_dimensions[row].height = 20

def data(ws, row, values, bg):
    for col, v in enumerate(values, 1):
        c = ws.cell(row, col, v)
        style(c, bg, sz=9, h='left')
    ws.row_dimensions[row].height = 36

# ─── DATA ────────────────────────────────────────────────────────────────────
# Cols: Genre,Prénom,Nom,Flag,Team,Vélo,Fourche,Amort,Freins,Trans,Roues,Pneus,Casque,Masque,Cockpit,Notes,Source

MEN = [
    ('M','Martin','MAES','🇧🇪','Orbea FMD Racing',
     'Orbea (DH proto)','Fox 38 Factory','Fox DHX2','Shimano XTR + Galfer rotors',
     'Shimano','DT Swiss 350 (mullet 29/27.5)','Maxxis Assegai 29" (F) + DHRII 27.5" (R)',
     'Met/Bluegrass','–','Race Face','Mullet setup · debut DH full time 2026',
     'Pinkbike / Vital MTB'),

    ('M','Max','ALRAN','🇫🇷','Commencal/Muc-Off by Riding Addiction',
     'Commencal Supreme DH V5','Fox 40 Factory','Fox Float X2 Factory',
     'Shimano','Shimano','Crankbrothers Synthesis','Schwalbe Magic Mary (F) + Tacky Chan (R)',
     '–','–','–','Junior WC champion 2025 → Elite 2026',
     'Commencal.com / Pinkbike'),

    ('M','Andreas','KOLB','🇦🇹','Santa Cruz Syndicate',
     'Santa Cruz V10 XL (custom MX dropout)','Fox 40 Factory (93-98 PSI, 4 tokens)','Fox Float X2 Factory',
     'Shimano XTR (F) + proto Saint caliper (R) 220mm','Shimano','Reserve 30 HD|AL (MX)',
     'Maxxis Assegai DH MaxxGrip 23 PSI (F) + DHR II DH MaxxGrip 28 PSI (R)',
     'Poc','Oakley','OneUp V2 Carbon 35mm rise 780mm + Wolf Tooth bar ends',
     'XL frame custom MX dropout · prototype Shimano rear caliper · OChain 12°',
     'Pinkbike Loudenvielle 2026'),

    ('M','Evan','MEDCALF','🇺🇸','Commencal Schwalbe by Les Orres',
     'Commencal Supreme DH','Fox 40 Factory','Fox Float X2 Factory',
     'SRAM','SRAM','Crankbrothers Synthesis','Schwalbe Magic Mary (F) + Tacky Chan (R)',
     'Troy Lee','–','–','Vice US Champion 2025 · 1er top 15 WC',
     'Commencal.com'),

    ('M','Aaron','GWIN','🇺🇸','Frameworks Racing / TRP',
     'Frameworks','Fox 40 Factory','Fox Float X2 Factory',
     'TRP','5DEV cranks + TRP','ENVE','Continental Kryptotal',
     'Bell','Oakley','Burgtec cockpit','Crankbrothers pedals · KMC chain · nouveau pour Gwin : Continental & Burgtec',
     'TRP Cycling / Bikerumor 2025'),

    ('M','Roger','VIEIRA','🇧🇷','Pivot Factory Racing',
     'Pivot Phoenix (size S1)','Fox 40 (70 PSI, 2 vol spacers)','Fox Float X2 (240 PSI)',
     'Shimano XTR 203mm F&R','Shimano Saint 10-speed + Saint cranks',
     'Reynolds Blacklabel DH 287r (MX)','Continental Argotal DH SuperSoft (22 PSI F · 26 PSI R)',
     'Troy Lee','Oakley','Renthal FatBar 35mm 30mm rise 770mm + ODI Longneck + Renthal stem 45mm',
     'Crankbrothers Mallet DH pédales · WTB Solano saddle',
     'Pinkbike Bielsko-Biala 2025'),

    ('M','Richard','RUDE JR','🇺🇸','Yeti / Fox Factory Race Team',
     'Yeti (Special Projects DH proto)','Fox 40 Factory','Fox DHX2 Factory',
     'Shimano','Shimano','DT Swiss','Maxxis Assegai 29" (F) + Minion DHR II (R)',
     'Kali','Oakley','–','Retour DH après carrière EWS · 4x EWS champion',
     'Yeti Cycles / Pinkbike 2025'),

    ('M','George','MADLEY','🇬🇧','Continental Atherton',
     'Atherton A.200','Fox 40 Factory GRIP2','Fox DHX2 Factory',
     'Continental / SRAM','SRAM X0 DH 1x7','STANS Flow EX3',
     'Continental Kryptotal SuperSoft (F) + Soft (R) DH casing',
     '–','–','–','3x Junior British Champion · proto World Cup setup',
     'Atherton Bikes / Vital MTB 2025'),

    ('M','Loïc','BRUNI','🇫🇷','Specialized Gravity',
     'Specialized Demo (proto custom)','Öhlins proto électronique','Öhlins',
     'Magura MT7','SRAM X01 DH','Roval (custom)','Maxxis (Assegai + Minion DHR II)',
     'Poc','Oakley','–','Mullet 29+27.5 · 3D-printed brake levers · suspension électronique prototype',
     'Bikerumor / Red Bull / Pinkbike 2025'),

    ('M','Luke','WALKER','🇦🇺','Privateer',
     '–','–','–','–','–','–','–','–','–','–',
     'Privateer australien · données limitées disponibles',
     'À compléter'),

    ('M','Bernard','KERR','🇬🇧','Pivot Factory Racing',
     'Pivot Phoenix','Fox 40 Factory','Fox Float X2 Factory',
     'SRAM Maven','SRAM GX DH','–','Maxxis',
     'Troy Lee','Oakley','–','3x Red Bull Hardline Winner · co-fondateur Pivot Factory Racing',
     'Pivot Cycles / Red Bull'),
]

WOMEN = [
    ('F','Valentina','ROA SANCHEZ','🇨🇴','MS Racing',
     'Zerode G3 (custom CNC)','SR Suntour Rux (modified)','Mara Piggy Back air shock',
     'Hope Tech (TR4 F · GR4 R)','Pinion Ci6 Gearbox + Gates belt drive',
     'Reynolds Blackline carbon','–',
     '–','–','Hope cockpit','High-pivot · belt drive · beaucoup de pièces Hope custom CNC · BYB data acq.',
     'Pinkbike / Bikerumor 2025'),

    ('F','Lisa','BOULADOU','🇫🇷','Santa Cruz / Burgtec by Goodman',
     'Santa Cruz V10','Fox 40 Factory','Fox Float X2 Factory',
     'SRAM Maven','SRAM','Burgtec wheels','Maxxis Assegai (F) + Minion DHR II (R)',
     '–','–','Burgtec cockpit','Junior WC win Leogang 2024 · Championne de France 2024',
     'Santa Cruz / Burgtec team pages'),

    ('F','Tahnée','SEAGRAVE','🇬🇧','Orbea FMD Racing',
     'Orbea Rallon DH (1er DH Orbea)','Fox 40 Factory GRIP X2 200mm','Fox Float X2 Factory 200mm',
     'Shimano XTR 200mm F&R','Shimano Saint 1x7',
     'OQUO MC32-Team (MX 29+27.5)','Maxxis Assegai MaxxGrip DH (F) + Minion DHR II MaxxGrip DH 2.4" (R)',
     'Fox Racing','Oakley','–','Bolt-on weights près du BB pour équilibre par piste · 1ère victoire WC avant le lancement officiel du vélo',
     'Orbea Stories / Bikerumor 2025'),

    ('F','Gracey','HEMSTREET','🇨🇦','Norco Race Division',
     'Norco prototype DH','Fox 40 (RAD, 78 PSI, 2 vol reducers)','Fox DHX2 (475 lb spring)',
     'Shimano XTR 203mm F&R','Shimano Saint 36T chainring',
     'Crankbrothers Synthesis Carbon DH MX + Chris King hubs',
     'Maxxis Assegai 24 PSI (F) + Minion DHR II 26 PSI (R)',
     '–','–','Race Face Atlas 35 · 50mm stem · 20mm rise · 750mm',
     '3x Red Bull Hardline women · WC Loudenvielle + Leogang + Les Gets 2025',
     'Pinkbike Hardline Tasmania 2024'),

    ('F','Harriet','HARNDEN','🇬🇧','AON Racing',
     'Sego (choix personnel, pas le Gamux)','Manitou Dorado inversée','Mara Piggy Back air shock',
     'Hayes Dominion (violet)','–','Reynolds Blackline carbon (custom)','Maxxis',
     '–','–','Hope cockpit · stem 10mm inversé · guidon 50mm rise',
     '77 victoires · EWS Overall 2024 · stem inversé signature',
     'Pinkbike / Bikerumor / Wideopen 2025'),

    ('F','Jess','BLEWITT','🇳🇿','Scott DH Factory Racing',
     'Scott Gambler RC','Fox 40 Factory (orange)','Fox Float X2 Factory',
     'Shimano','Shimano','DT Swiss','Michelin (sponsor Scott DH)',
     '–','–','–','2x NZ National Champion · 1ère femme Red Bull Hardline',
     'Pinkbike / Scott DH Factory 2025'),

    ('F','Frida Helena','RØNNING','🇳🇴','Merida NTG',
     'Merida One-Sixty (DH)','Fox 40','Fox Float X2',
     'Shimano','Shimano','–','Maxxis',
     '–','–','–','3x Championne de Norvège · PhD Mech Engineering UTK',
     'À compléter'),

    ('F','Laïs','BONNAURE','🇫🇷','VTT Luberon',
     '–','Fox 36','Fox Float X2',
     'SRAM','SRAM','–','Maxxis',
     '–','–','–','Championne de France 2022 & 2024 · Championne Europe U23 2023',
     'À compléter'),
]

row = 3
section(ws, row, '— MEN ELITE —', NCOLS, C_BG, C_LIME)
row += 1
for i, d in enumerate(MEN):
    bg = C_MEN_A if i % 2 == 0 else C_MEN_B
    data(ws, row, d, bg)
    row += 1

section(ws, row, '— WOMEN ELITE —', NCOLS, C_PURPLE, C_LIME)
row += 1
for i, d in enumerate(WOMEN):
    bg = C_WOM_A if i % 2 == 0 else C_WOM_B
    data(ws, row, d, bg)
    row += 1

# Freeze top 2 rows
ws.freeze_panes = 'A3'

wb.save('/sessions/dreamy-vibrant-volta/mnt/freeride/Equipements_Riders_2026.xlsx')
print(f"✅ Sauvegardé — {row-4} riders ({len(MEN)} M + {len(WOMEN)} F)")
