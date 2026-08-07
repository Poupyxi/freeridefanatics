#!/usr/bin/env python3
"""
FreerideFanatics — Google Sheet sync.

Downloads the UCI_DH_2026_Tracker_v3 Google Sheet (public link-share export)
and regenerates data/riders.json from its tabs:

  - "🔧 Equipment Women" / "🔧 Equipment Men"  → per-rider equipment (keyed by Instagram)
  - "👤 Profils"                                → identity, team, sponsors, bio
  - "📊 Résultats 2026"                         → country code + UCI points per round
  - "equipment link"                            → product and Amazon URLs per part (shop buttons)
                                                  + optional Photo URL column (downloaded
                                                  into assets/img/equipment/ when filled)

Then rebuilds the whole site (same as running build.py).

Usage:
  python3 sync.py                  # sync + rebuild
  python3 sync.py --no-build       # only refresh data/riders.json
  python3 sync.py --offline FILE   # parse a local .xlsx instead of downloading
"""
import json
import io
import os
import re
import ssl
import subprocess
import sys
import unicodedata
import urllib.request

import openpyxl
from PIL import Image, ImageOps

import build  # equip_image_slug + site generator

SHEET_ID = "1xyD72CIBG4TQLFmvhNRVKfyz2ZsKn3HUFA9aXWZH9Lc"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

ROOT = os.path.dirname(os.path.abspath(__file__))
DATA_PATH = os.path.join(ROOT, "data", "riders.json")
EQUIP_IMG_DIR = os.path.join(ROOT, "assets", "img", "equipment")

SEASON = 2026
DISCIPLINE = "UCI MTB World Cup Downhill (DH)"
HISTORY_CATEGORY = "UCI MTB World Cup DH 2026"

TAB_EQUIP_WOMEN = "Equipment Women"
TAB_EQUIP_MEN = "Equipment Men"
TAB_PROFILES = "Profils"
TAB_RESULTS = ("Résultats 2026", "UCI World CUP 2026")
TAB_LINKS = "equipment link"

# Column headers in the equipment tabs → category keys used in riders.json
CATEGORY_KEYS = {
    "Frame": "Frame", "Fork": "Fork", "Rear Shock": "RearShock",
    "Handlebar": "Handlebar", "Dropper Post": "DropperPost", "Saddle": "Saddle",
    "Crankset": "Crankset", "Derailleur": "Derailleur", "Brake Lever": "BrakeLever",
    "GRIP": "GRIP", "CHAIN": "CHAIN", "Disk": "Disk",
    "Wheels": "Wheels", "Tires": "Tires", "Pedals": "Pedals",
    "Shoes": "Shoes", "Helmet": "Helmet", "Protection": "Protection", "Goggles": "Goggles",
}

# French → English fixes for event column headers
EVENT_TRANSLATIONS = {
    "autriche": "Austria", "italie": "Italy", "suisse": "Switzerland",
}
# The sheet mixes French and English country names; the site is in English.
COUNTRY_TRANSLATIONS = {
    "états-unis": "USA", "etats-unis": "USA", "united states": "USA",
    "grande-bretagne": "Great Britain", "royaume-uni": "Great Britain",
    "belgique": "Belgium", "brésil": "Brazil", "bresil": "Brazil",
    "norwey": "Norway", "norvège": "Norway", "norvege": "Norway",
    "autriche": "Austria", "suisse": "Switzerland", "italie": "Italy",
    "espagne": "Spain", "allemagne": "Germany", "irlande": "Ireland",
    "nouvelle-zélande": "New Zealand", "nouvelle-zelande": "New Zealand",
    "afrique du sud": "South Africa", "pologne": "Poland",
    "slovénie": "Slovenia", "slovenie": "Slovenia", "tchéquie": "Czech Republic",
}
MONTH_TRANSLATIONS = {
    "may": "May", "mai": "May", "june": "June", "juin": "June", "july": "July",
    "juillet": "July", "august": "August", "aout": "August", "août": "August",
    "september": "September", "sept.": "September", "sept": "September",
    "septembre": "September", "october": "October", "octobre": "October",
}

# ---------------------------------------------------------------- helpers

def _ssl_context():
    try:
        import certifi
        return ssl.create_default_context(cafile=certifi.where())
    except ImportError:
        return ssl.create_default_context()

def fetch(url, timeout=60):
    """Download url; falls back to curl when Python has no usable CA bundle."""
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urllib.request.urlopen(req, timeout=timeout, context=_ssl_context()) as resp:
            return resp.read()
    except Exception:
        out = subprocess.run(["curl", "-sfL", "--max-time", str(timeout), url],
                             capture_output=True)
        if out.returncode != 0:
            raise
        return out.stdout

def find_sheet(wb, name_part):
    candidates = (name_part,) if isinstance(name_part, str) else name_part
    for ws in wb.worksheets:
        if any(candidate.lower() in ws.title.lower() for candidate in candidates):
            return ws
    expected = "' or '".join(candidates)
    raise SystemExit(f"Tab matching '{expected}' not found — sheet layout changed?")

def clean(v):
    if v is None:
        return ""
    s = str(v).replace(" ", " ").strip()
    return s

def strip_emoji(s):
    return "".join(c for c in s if unicodedata.category(c) not in ("So", "Sk", "Cs", "Cf")).strip()

def slugify(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")
    return s

def as_int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None

def normalize_event(header):
    """'🇦🇹 Autriche\\n(june)' -> 'Austria (June)'"""
    s = strip_emoji(clean(header)).replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip()
    m = re.match(r"^(.*?)\s*\(([^)]+)\)\s*$", s)
    if not m:
        return s
    place, month = m.group(1).strip(), m.group(2).strip()
    place = EVENT_TRANSLATIONS.get(place.lower(), place)
    month = MONTH_TRANSLATIONS.get(month.lower(), month.capitalize())
    return f"{place} ({month})"

def normalize_bio(s):
    s = clean(s).replace("×", "x").replace("\n", " ")
    return re.sub(r"\s+", " ", s).strip()

def normalize_country(s):
    return COUNTRY_TRANSLATIONS.get(s.lower(), s) if s else None

# Placeholders the sheet uses for "no team"; kept as-is they surface as a real
# team called "N/A" in the standings instead of falling back to Privateer.
NO_VALUE = {"n/a", "na", "n.a.", "-", "--", "—", "?", "none", "tbd", "n/c"}

def normalize_team(s):
    return None if not s or s.strip().lower() in NO_VALUE else s.strip()

def section_gender(cells, current):
    """Track the 'WOMEN ELITE' / 'MEN ELITE' section header rows.
    Some rider rows have a dirty G cell (stray newlines), so the section is
    the reliable fallback."""
    joined = " ".join(cells).upper()
    if "WOMEN ELITE" in joined:
        return "W"
    if "MEN ELITE" in joined:
        return "M"
    return current

# ---------------------------------------------------------------- parsers

def parse_profiles(ws):
    """Rows keyed by instagram handle; also returns sheet order."""
    profiles = {}
    header_seen = False
    gender = None
    for row in ws.iter_rows(values_only=True):
        cells = [clean(c) for c in row]
        if not header_seen:
            if cells[0] == "G":
                header_seen = True
            continue
        gender = section_gender(cells, gender)
        g, first, last = cells[0], cells[1], cells[2]
        if g not in ("W", "M"):
            g = gender
        if g not in ("W", "M") or not first or not last:
            continue  # section headers / blanks
        insta = cells[8]
        profiles[insta] = {
            "gender_category": "Women Elite" if g == "W" else "Men Elite",
            "first_name": first,
            "last_name": last,
            "country": normalize_country(cells[4]),
            "hometown": cells[5] or None,
            "date_of_birth": cells[6] or None,
            "age": as_int(cells[7]),
            "instagram": insta or None,
            "bio": normalize_bio(cells[9]),
            "team": normalize_team(cells[10]),
            "sponsors": [p.strip() for p in cells[11].split(";") if p.strip()] if cells[11] else [],
        }
    return profiles

def parse_equipment_tab(ws):
    """instagram -> list of raw (category_key, cell_value)."""
    out = {}
    categories = None
    for row in ws.iter_rows(values_only=True):
        cells = [clean(c) for c in row]
        if categories is None:
            if cells[0] == "G":
                categories = [(i, CATEGORY_KEYS[h]) for i, h in enumerate(cells) if h in CATEGORY_KEYS]
            continue
        if not cells[1] or not cells[1].startswith("@"):
            continue
        insta = cells[1]
        items = []
        for col, cat in categories:
            if col < len(cells) and cells[col]:
                items.append((cat, cells[col]))
        out[insta] = items
    return out

# Non-finishing statuses the sheet writes in the Classement cell.
DNX_STATUSES = {"DNS", "DNF", "DSQ", "DQ", "DNQ", "OTL"}

def parse_results(ws):
    """(first, LAST) lowercase -> {'country_code': str, 'history': [...]}.

    Each round occupies TWO columns under one merged title:

        | 🇦🇹 Autriche-Leogang (june) |          <- row 2, merged across both
        | Classement    | Points      |          <- row 3

    openpyxl exposes a merged range's value on its top-left cell only, so the
    column index recovered from the title row is the *Classement* (finishing
    position) column and the points sit one column to its right. Reading the
    title column as points is what inverted the whole standings: positions were
    summed as if they were points, so the slowest riders ranked highest.
    Positions are kept as-is from the sheet and stored separately."""
    out = {}
    events = None
    for row in ws.iter_rows(values_only=True):
        cells = [clean(c) for c in row]
        if events is None:
            if cells[0] == "G":
                events = [(i, i + 1, normalize_event(h))
                          for i, h in enumerate(cells[5:], start=5) if strip_emoji(h)]
            continue
        if not cells[1] or not cells[2] or "ELITE" in cells[0].upper():
            continue
        history = []
        for rank_col, points_col, event in events:
            rank_raw = cells[rank_col] if rank_col < len(cells) else ""
            points_raw = cells[points_col] if points_col < len(cells) else ""
            pts = as_int(points_raw)
            place = as_int(rank_raw)
            status = rank_raw.upper() if rank_raw.upper() in DNX_STATUSES else None
            if pts is None and place is None and status is None:
                continue  # rider did not contest this round
            history.append({
                "year": SEASON, "event": event, "category": HISTORY_CATEGORY,
                "result": status or build.ordinal(place),
                "place": None if status else place,
                "points": pts,
            })
        out[(slugify(cells[1]), slugify(cells[2]))] = {
            "country_code": cells[3] or None,
            "history": history,
        }
    return out

def norm_detail(s):
    """Loose comparison form: 'Spike 35 Vibrocore™ Bar 25mm' -> 'spike 35 vibrocore bar 25mm'."""
    s = s.lower().replace("™", "").replace("®", "")
    return re.sub(r"[^a-z0-9+]+", " ", s).strip()

def parse_links(ws):
    """Product catalogue: {'exact': {(cat, brand, detail): entry}, 'index': {(cat, brand): [entries]}}."""
    exact = {}
    index = {}
    category = None
    headers = [norm_detail(clean(cell.value)) for cell in ws[1]]

    def column_index(*names):
        for name in names:
            normalized = norm_detail(name)
            if normalized in headers:
                return headers.index(normalized)
        return None

    link_col = column_index("Link", "Product link", "Lien", "Lien produit")
    amazon_col = column_index("Amazon", "Amazon FR COM", "Amazon link", "Lien Amazon")
    photo_col = column_index("Photo", "Photo URL", "Image", "Image URL")
    # Backward compatibility with the original three-column catalogue.
    link_col = 2 if link_col is None else link_col
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = [clean(c) for c in row]
        brand = cells[0] if cells else ""
        detail = cells[1] if len(cells) > 1 else ""
        link = cells[link_col] if link_col < len(cells) else ""
        amazon_link = cells[amazon_col] if amazon_col is not None and amazon_col < len(cells) else ""
        photo = cells[photo_col] if photo_col is not None and photo_col < len(cells) else ""
        if not brand:
            continue
        if not detail and not link:  # section header row ("Frame", "Fork", …)
            key = CATEGORY_KEYS.get(brand) or CATEGORY_KEYS.get(brand.title())
            if key:
                category = key
            continue
        if category is None:
            continue
        detail_parts = [part.strip() for part in detail.split(";")]
        canonical_brand, canonical_main = build.canonical_equipment_product(
            category, brand, detail_parts[0] if detail_parts else "")
        canonical_detail = ";".join([canonical_main] + detail_parts[1:])
        entry = {
            "detail": norm_detail(canonical_detail),
            "main": norm_detail(canonical_main),
            "link": link or None,
            "amazon_link": amazon_link or None,
            "photo": photo or None,
            "aliased": (build.norm_product_text(brand) != build.norm_product_text(canonical_brand)
                        or build.norm_product_text(detail_parts[0] if detail_parts else "")
                        != build.norm_product_text(canonical_main)),
        }
        brand_key = build.norm_product_text(canonical_brand)
        exact[(category, brand_key, norm_detail(canonical_detail))] = entry
        index.setdefault((category, brand_key), []).append(entry)
    return {"exact": exact, "index": index}

def resolve_link(links, cat, brand, model_detail):
    """Best product entry for an equipment cell — exact detail first, then the
    main model name, then the longest normalized prefix overlap."""
    parts = [part.strip() for part in model_detail.split(";")]
    canonical_brand, canonical_main = build.canonical_equipment_product(
        cat, brand, parts[0] if parts else "")
    canonical_detail = ";".join([canonical_main] + parts[1:])
    brand_key = build.norm_product_text(canonical_brand)
    candidates = links["index"].get((cat, brand_key))
    if not candidates:
        return None
    main = norm_detail(canonical_main)
    if not main:
        return None
    entry = links["exact"].get((cat, brand_key, norm_detail(canonical_detail)))
    if entry:
        # If this catalogue row used an alias (V5), prefer the canonical row's
        # current URL/photo (V5.2) so one product cannot retain two destinations.
        preferred = [e for e in candidates if e["main"] == main and not e.get("aliased")
                     and (e.get("link") or e.get("photo"))]
        return preferred[0] if entry.get("aliased") and preferred else entry
    same_main = [e for e in candidates if e["main"] == main]
    if same_main:
        canonical_rows = [e for e in same_main if not e.get("aliased")]
        return (canonical_rows or same_main)[0]
    prefix = [e for e in candidates
              if e["main"].startswith(main + " ") or main.startswith(e["main"] + " ")]
    if prefix:
        # most specific name wins (longest overlap), stable on ties
        return max(prefix, key=lambda e: min(len(e["main"]), len(main)))
    return None

# ---------------------------------------------------------------- assembly

def build_riders(wb):
    profiles = parse_profiles(find_sheet(wb, TAB_PROFILES))
    equipment = {}
    equipment.update(parse_equipment_tab(find_sheet(wb, TAB_EQUIP_WOMEN)))
    equipment.update(parse_equipment_tab(find_sheet(wb, TAB_EQUIP_MEN)))
    results = parse_results(find_sheet(wb, TAB_RESULTS))
    links = parse_links(find_sheet(wb, TAB_LINKS))

    unmatched_links = 0
    photos = {}  # equip slug -> photo url
    riders = []
    for insta, p in profiles.items():
        first, last = p["first_name"], p["last_name"]
        res = results.get((slugify(first), slugify(last)), {})

        items = []
        bike = None
        for cat, raw in equipment.get(insta, []):
            parts = [x.strip() for x in raw.split(";")]
            raw_brand = parts[0]
            raw_detail_parts = parts[1:]
            raw_main = raw_detail_parts[0] if raw_detail_parts else ""
            brand, canonical_main = build.canonical_equipment_product(cat, raw_brand, raw_main)
            model_detail = ";".join([canonical_main] + raw_detail_parts[1:]) if raw_detail_parts else ""
            entry = resolve_link(links, cat, brand, model_detail) if model_detail else None
            if entry is None and model_detail:
                unmatched_links += 1
            link = entry["link"] if entry else None
            amazon_link = entry["amazon_link"] if entry else None
            if entry and entry["photo"]:
                photos[build.equip_image_slug(cat, brand, model_detail.split(";")[0])] = entry["photo"]
            items.append({
                "category": cat,
                "brand": brand,
                "model_detail": model_detail,
                "brand_model": ";".join([brand, model_detail]) if model_detail else brand,
                "affiliate_link": link,
                "amazon_link": amazon_link,
            })
            if cat == "Frame" and bike is None:
                bike = {"brand": brand, "model": model_detail}

        riders.append({
            "name": f"{first} {last}",
            "first_name": first,
            "last_name": last,
            "gender_category": p["gender_category"],
            "discipline": DISCIPLINE,
            "country": p["country"],
            "country_code": res.get("country_code"),
            "hometown": p["hometown"],
            "date_of_birth": p["date_of_birth"],
            "age": p["age"],
            "instagram": p["instagram"],
            "team": p["team"],
            "sponsors": p["sponsors"],
            "bio": p["bio"],
            "bike": bike,
            "equipment": items,
            "competition_history": res.get("history", []),
            "photo_url": None,
            "season": SEASON,
            "slug": slugify(f"{first} {last}"),
            "display_name": f"{first} {last.title()}",
        })

    riders.sort(key=lambda r: (r["last_name"].lower(), r["first_name"].lower()))
    return riders, photos, unmatched_links

def download_photos(photos):
    """Fetch equipment photo URLs from the sheet into assets/img/equipment/."""
    os.makedirs(EQUIP_IMG_DIR, exist_ok=True)
    fetched = 0
    for slug, url in photos.items():
        if not url.startswith("http"):
            continue
        existing = [os.path.join(EQUIP_IMG_DIR, f"{slug}.{e}")
                    for e in ("jpg", "jpeg", "png", "webp")
                    if os.path.exists(os.path.join(EQUIP_IMG_DIR, f"{slug}.{e}"))]
        valid_existing = False
        for path in existing:
            try:
                with Image.open(path) as image:
                    image.verify()
                valid_existing = True
            except (OSError, ValueError):
                os.remove(path)
                print(f"  ! removed invalid photo for {slug}")
        if valid_existing:
            continue
        try:
            data = fetch(url, timeout=20)
            with Image.open(io.BytesIO(data)) as source:
                image = ImageOps.exif_transpose(source).convert("RGB")
                image.thumbnail((600, 600), Image.LANCZOS)
                image.save(os.path.join(EQUIP_IMG_DIR, f"{slug}.jpg"), "JPEG",
                           quality=82, optimize=True, progressive=True)
            fetched += 1
        except (OSError, ValueError) as e:
            print(f"  ! photo failed for {slug}: {e}")
    return fetched

# ---------------------------------------------------------------- main

def main():
    args = sys.argv[1:]
    do_build = "--no-build" not in args

    if "--offline" in args:
        path = args[args.index("--offline") + 1]
    else:
        path = os.path.join(ROOT, ".sheet-cache.xlsx")
        print(f"Downloading sheet {SHEET_ID}…")
        data = fetch(EXPORT_URL)
        with open(path, "wb") as f:
            f.write(data)

    wb = openpyxl.load_workbook(path, data_only=True)
    riders, photos, unmatched = build_riders(wb)

    women = sum(1 for r in riders if r["gender_category"] == "Women Elite")
    linked = sum(1 for r in riders for i in r["equipment"] if i["affiliate_link"])
    amazon_linked = sum(1 for r in riders for i in r["equipment"] if i["amazon_link"])
    total_items = sum(len(r["equipment"]) for r in riders)

    with open(DATA_PATH, "w", encoding="utf-8") as f:
        json.dump(riders, f, ensure_ascii=False, indent=1)
        f.write("\n")

    print(f"riders.json: {len(riders)} riders ({women} women, {len(riders) - women} men)")
    print(f"equipment:   {total_items} items, {linked} with product link, "
          f"{amazon_linked} with Amazon link" +
          (f", {unmatched} without a match in 'equipment link' tab" if unmatched else ""))

    if photos:
        n = download_photos(photos)
        print(f"photos:      {len(photos)} URL(s) in sheet, {n} downloaded")

    if do_build:
        build.main()

if __name__ == "__main__":
    main()
