from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C_BG    = '1A1A2E'
C_LIME  = 'C8D400'
C_WHITE = 'FFFFFF'
C_MEN_A = 'F5F5F5'
C_MEN_B = 'FFFFFF'
C_WOM_A = 'FFF0F5'
C_WOM_B = 'FFF8FC'
C_SEP_M = '1A1A2E'
C_SEP_W = '4A1942'

COLS = [
    'Genre', 'Prénom', 'Nom', 'Flag', 'Team',
    'Vélo / Cadre', 'Fourche', 'Amortisseur',
    'Freins', 'Transmission', 'Roues', 'Pneus',
    'Casque', 'Masque', 'Cockpit / Guidon',
    'Notes', 'Source'
]
COL_WIDTHS = [7, 14, 18, 6, 34, 24, 22, 22, 20, 26, 30, 36, 16, 16, 22, 28, 32]

# ── DATA ────────────────────────────────────────────────────────────────────
# (Genre, Prénom, Nom, Flag, Team, Vélo, Fourche, Amortisseur, Freins,
#  Transmission, Roues, Pneus, Casque, Masque, Cockpit, Notes, Source)

MEN = [
    ('M','Loïc','BRUNI','🇫🇷','Specialized Gravity',
     'Specialized Demo 11','Öhlins RXF 38 (proto)','Öhlins TTX22 (proto)',
     'SRAM Maven Ultimate','SRAM XX DH AXS T-Type','Roval Traverse Gravity (DT Swiss 350)','Maxxis Assegai 29 (F) + Minion DHR II 27.5 (R)',
     'Poc','Poc / Oakley','Specialized / Burgtec','Proto Öhlins fork & shock (Bruni only)','Pinkbike / Specialized'),

    ('M','Finn','ILES','🇨🇦','Specialized Gravity',
     'Specialized Demo 11','Fox 40 Factory GRIP X2','Fox DHX2 Factory coil',
     'SRAM Maven Ultimate','SRAM XX DH AXS T-Type','Roval Traverse Gravity (DT Swiss 350)','Maxxis Assegai 29 (F) + Minion DHR II 27.5 (R)',
     'Poc','Poc / Oakley','Specialized / Burgtec','Mullet 29/27.5','Pinkbike / Specialized'),

    ('M','Jordan','WILLIAMS','🇬🇧','Specialized Gravity',
     'Specialized Demo 11','Fox 40 Factory GRIP X2','Fox DHX2 Factory coil',
     'SRAM Maven Ultimate','SRAM XX DH AXS T-Type','Roval Traverse Gravity (DT Swiss 350)','Maxxis Assegai 29 (F) + Minion DHR II 27.5 (R)',
     'Poc','Poc / Oakley','Specialized / Burgtec','Mullet 29/27.5','Pinkbike / Specialized'),

    ('M','Jackson','GOLDSTONE','🇨🇦','Santa Cruz Syndicate',
     'Santa Cruz V10','Fox 40 Factory GRIP X2','Fox Float X2 Factory',
     'Shimano Saint','SRAM X01 DH','Reserve 30|HD carbon','Maxxis Assegai (F) + Minion DHR II (R)',
     'Bell','Bell / Oakley','Race Face','','Pinkbike / Santa Cruz'),

    ('M','Andreas','KOLB','🇦🇹','Santa Cruz Syndicate',
     'Santa Cruz V10','Fox 40 Factory GRIP X2','Fox Float X2 Factory',
     'Shimano Saint','SRAM X01 DH','Reserve 30|HD carbon','Maxxis Assegai (F) + Minion DHR II (R)',
     'Bell','Bell / Oakley','Race Face','','Pinkbike / Santa Cruz'),

    ('M','Amaury','PIERRON','🇫🇷','Commencal / Muc-Off by Riding Addiction',
     'Commencal Supreme DH V5','Fox 40 Factory GRIP X2','Fox Float X2 Factory',
     'SRAM Maven','SRAM X01 DH','Crankbrothers Synthesis DH','Schwalbe Magic Mary (F) + Tacky Chan (R)',
     'Leatt','Leatt / Oakley','Race Face','','Pinkbike / Commencal'),

    ('M','Loris','VERGIER','🇫🇷','Commencal / Muc-Off by Riding Addiction',
     'Commencal Supreme DH V5','Fox 40 Factory GRIP X2','Fox Float X2 Factory',
     'SRAM Maven','SRAM X01 DH','Crankbrothers Synthesis DH','Schwalbe Magic Mary (F) + Tacky Chan (R)',
     'Leatt','Leatt / Oakley','Race Face','','Pinkbike / Commencal'),

    ('M','Till','ALRAN','🇫🇷','Commencal / Muc-Off by Riding Addiction',
     'Commencal Supreme DH V5','Fox 40 Factory GRIP X2','Fox Float X2 Factory',
     'SRAM Maven','SRAM X01 DH','Crankbrothers Synthesis DH','Schwalbe Magic Mary (F) + Tacky Chan (R)',
     'Leatt','Leatt / Oakley','Race Face','','Pinkbike / Commencal'),

    ('M','Max','ALRAN','🇫🇷','Commencal / Muc-Off by Riding Addiction',
     'Commencal Supreme DH V5','Fox 40 Factory GRIP X2','Fox Float X2 Factory',
     'SRAM Maven','SRAM X01 DH','Crankbrothers Synthesis DH','Schwalbe Magic Mary (F) + Tacky Chan (R)',
     'Leatt','Leatt / Oakley','Race Face','','Pinkbike / Commencal'),

    ('M','Antoine','PIERRON','🇫🇷','Commencal Schwalbe by Les Orres',
     'Commencal Supreme DH V5','Fox 40 Factory GRIP X2','Fox Float X2 Factory',
     'SRAM Maven Silver','SRAM X01 DH','Crankbrothers Synthesis DH','Schwalbe Magic Mary (F) + Hans Dampf (R)',
     'Leatt','Leatt / Oakley','Race Face','Les Orres team = Schwalbe tires','Pinkbike / Commencal'),

    ('M','Evan','MEDCALF','🇬🇧','Commencal Schwalbe by Les Orres',
     'Commencal Supreme DH V5','Fox 40 Factory GRIP X2','Fox Float X2 Factory',
     'SRAM Maven Silver','SRAM X01 DH','Crankbrothers Synthesis DH','Schwalbe Magic Mary (F) + Hans Dampf (R)',
     'Leatt','Leatt / Oakley','Race Face','','Pinkbike / Commencal'),

    ('M','Troy','BROSNAN','🇦🇺','Canyon DH Racing',
     'Canyon Sender CFR (high-pivot carbon)','RockShox Boxxer Ultimate','RockShox Vivid Coil Ultimate',
     'SRAM Maven Silver','SRAM X01 DH','DT Swiss EX 1700 Spline MX','Maxxis Assegai (F) + Minion DHR II (R)',
     'Troy Lee Designs','Troy Lee Designs / Oakley','Race Face','Mullet option disponible','Pinkbike / Canyon'),

    ('M','Luca','SHAW','🇺🇸','Canyon DH Racing',
     'Canyon Sender CFR (high-pivot carbon)','RockShox Boxxer Ultimate','RockShox Vivid Coil Ultimate',
     'SRAM Maven Silver','SRAM X01 DH','DT Swiss EX 1700 Spline MX','Maxxis Assegai (F) + High Roller II (R)',
     'Troy Lee Designs','Troy Lee Designs / Oakley','Race Face','Shaw utilise Flight Attendant sur certaines manches','Pinkbike / Canyon'),

    ('M','Henri','KIEFER','🇩🇪','Canyon DH Racing',
     'Canyon Sender CFR (high-pivot carbon)','RockShox Boxxer Ultimate','RockShox Vivid Coil Ultimate',
     'SRAM Maven Silver','SRAM X01 DH','DT Swiss EX 1700 Spline MX','Maxxis Assegai (F) + Minion DHR II (R)',
     'Troy Lee Designs','Troy Lee Designs / Oakley','Race Face','','Pinkbike / Canyon'),

    ('M','Ryan','PINKERTON','🇺🇸','Mondraker Factory Racing DH',
     'Mondraker Summum','RockShox Boxxer Ultimate','RockShox Vivid Coil Ultimate',
     'SRAM Maven','SRAM XX DH AXS T-Type','DT Swiss EX 1700 Spline MX','Maxxis Assegai (F) + Minion DHR II (R)',
     'Alpinestars','Oakley','Renthal 760mm','Champion USA 2025 — 3 ans contrat Mondraker','Pinkbike / Mondraker'),

    ('M','Asa','VERMETTE','🇨🇦','Frameworks Racing / TRP',
     'Frameworks carbon DH','Fox 40 Factory GRIP X2','Fox Factory coil',
     'TRP DH-EVO 4-piston 220mm','TRP / 5DEV cranks','ENVE carbon DH','Continental DHF (F) + DHR II (R)',
     'Troy Lee Designs','Troy Lee Designs / Oakley','Burgtec','','Pinkbike / Frameworks'),

    ('M','Aaron','GWIN','🇺🇸','Frameworks Racing / TRP',
     'Frameworks carbon DH','Fox 40 Factory GRIP X2','Fox Factory coil',
     'TRP DH-EVO 4-piston 220mm','TRP / 5DEV cranks','ENVE carbon DH','Continental DHF (F) + DHR II (R)',
     'Troy Lee Designs','Troy Lee Designs / Oakley','Burgtec','Retour compétition','Pinkbike / Frameworks'),

    ('M','Bodhi','KUHN','🇨🇭','Norco x Adidas Race Division',
     'Norco Aurum HSP prototype','Fox 40 RAD 203mm','Fox DHX2 coil',
     'Shimano XTR 4-piston 220mm','Shimano XTR 12sp','Crankbrothers Synthesis Carbon DH (Chris King hubs)','Maxxis Assegai (F) + Minion DHF (R)',
     'Poc','Poc / Adidas','OneUp Components','','Pinkbike / Norco'),

    ('M','Ethan','CRAIK','🇬🇧','Scott Downhill Factory',
     'Scott Gambler RC (6-bar carbon, 210mm)','Fox 40 Factory GRIP X2','Fox DHX2 Factory',
     'SRAM Maven Silver','SRAM X01 DH','DT Swiss EX 1700 Spline MX','Michelin DH22 (F) + Wild Enduro (R)',
     'Scott','Scott / Oakley','Syncros','','Pinkbike / Scott'),

    ('M','Dakotah','NORTON','🇺🇸','Scott Downhill Factory',
     'Scott Gambler RC (6-bar carbon, 210mm)','Fox 40 Factory GRIP X2','Fox DHX2 Factory',
     'SRAM Maven Silver','SRAM X01 DH','DT Swiss EX 1700 Spline MX','Michelin DH22 (F) + Wild Enduro (R)',
     'Scott','Scott / Oakley','Syncros','','Pinkbike / Scott'),

    ('M','Benoit','COULANGES','🇫🇷','Scott Downhill Factory',
     'Scott Gambler RC (proto rocker alu brut)','Fox 40 Factory GRIP X2','Fox DHX2 Factory',
     'SRAM Maven Silver','SRAM X01 DH','DT Swiss EX 1700 Spline MX','Michelin DH22 (F) + Wild Enduro (R)',
     'Scott','Scott / Oakley','Syncros','Rocker link aluminium brut prototype','Pinkbike / Scott'),

    ('M','Oliver','DAVIS','🇬🇧','Trek - Unbroken DH',
     'Trek Session DH','RockShox Boxxer Factory','RockShox Vivid Coil Factory',
     'SRAM Maven Silver','SRAM X01 DH','Bontrager Line DH','Maxxis Assegai (F) + Minion DHR II (R)',
     'Fox Racing','Fox Racing / Oakley','Burgtec','','Pinkbike / Trek'),

    ('M','Lachlan','STEVENS-MCNAB','🇦🇺','Trek - Unbroken DH',
     'Trek Session DH','RockShox Boxxer Factory','RockShox Vivid Coil Factory',
     'SRAM Maven Silver','SRAM X01 DH','Bontrager Line DH','Maxxis Assegai (F) + Minion DHR II (R)',
     'Fox Racing','Fox Racing / Oakley','Burgtec','','Pinkbike / Trek'),

    ('M','Luke','MEIER-SMITH','🇦🇺','Giant Factory Off-Road Team DH',
     'Giant Glory Advanced (mullet 29/27.5)','Fox 40 Factory GRIP2 203mm','Fox DHX2 coil 500lb',
     'Shimano Saint 220mm','Shimano Saint 7sp','Giant DH / Burgtec bar & stem','Maxxis Assegai 29 (F) + Minion DHR II 27.5 (R)',
     'Fox Racing','Fox Racing / Oakley','Burgtec 770mm 30mm rise','OChain 9° — tuned mass damper','Pinkbike bike check MSA 2025'),

    ('M','Charlie','HATTON','🇬🇧','Continental Atherton',
     'Atherton A.200','Fox 40 GRIP2 203mm','Fox DHX2 Factory coil',
     'SRAM Maven Silver','SRAM X0 DH 7sp','STANS Flow EX3 MX','Continental Kryptotal (F) + Trail King (R)',
     'Bell','Bell / Oakley','Renthal 780mm','','Pinkbike / Atherton'),

    ('M','Luke','WAYMAN','🇬🇧','Continental Atherton',
     'Atherton A.200','Fox 40 GRIP2 203mm','Fox DHX2 Factory coil',
     'SRAM Maven Silver','SRAM X0 DH 7sp','STANS Flow EX3 MX','Continental Kryptotal (F) + Trail King (R)',
     'Bell','Bell / Oakley','Renthal 780mm','','Pinkbike / Atherton'),

    ('M','Tuhoto-Ariki','PENE','🇳🇿','MS-Racing',
     'Zerode G3 (Pinion gearbox + belt)','SR Suntour Rux (custom damper)','Mara Piggy Back coil',
     'Hope Tech 4 (TR4 F + GR4 R) 203mm','Pinion C1.12 + Gates CDX belt','Reynolds Blackline DH carbon','Maxxis Assegai (F) + Minion DHR II (R)',
     'Bell','Bell / Oakley','Renthal 780mm','Gearbox + belt drive, pas de dérailleur','Pinkbike / MS-Racing'),

    ('M','Nathan','PONTVIANNE','🇫🇷','Santa Cruz Burgtec by Goodman',
     'Santa Cruz V10','Fox 40 Factory GRIP X2','Fox Float X2 Factory',
     'SRAM Maven','SRAM X01 DH','Burgtec EVO-D MK4','Maxxis Assegai (F) + Minion DHR II (R)',
     'Leatt','Leatt / Oakley','Race Face / Burgtec','','Pinkbike / Santa Cruz Burgtec'),

    ('M','Dylan','MAPLES','🇺🇸','Pivot Factory Racing',
     'Pivot Phoenix FF','Fox 40 Factory GRIP X2','Fox DHX2 Factory coil',
     'SRAM Maven','SRAM X01 DH','Burgtec EVO-D MK4','Maxxis Assegai (F) + Minion DHR II (R)',
     'Fox Racing','Fox Racing / Oakley','Race Face','','Pinkbike / Pivot'),

    ('M','Roger','VIEIRA','🇧🇷','Pivot Factory Racing',
     'Pivot Phoenix FF','Fox 40 Factory GRIP X2','Fox DHX2 Factory coil',
     'SRAM Maven','SRAM X01 DH','Burgtec EVO-D MK4','Maxxis Assegai (F) + Minion DHR II (R)',
     'Fox Racing','Fox Racing / Oakley','Race Face','','Pinkbike / Pivot'),

    ('M','Reece','WILSON','🇬🇧','AON Racing',
     'Gamux prototype (CNC alum.)','Manitou Dorado Expert 200mm','Mara coil',
     'Hayes Dominion A4','Pinion gearbox + Gates CDX belt','Reynolds Blackline DH carbon','Maxxis Assegai (F) + Minion DHR II (R)',
     'Bell','Bell / Oakley','Renthal 780mm','Gamux: gearbox + belt, Reynolds carbone, Hayes violet','Pinkbike / AON Racing'),

    ('M','Tyler','WAITE','🇺🇸','Yeti / Fox Factory Race Team',
     'Yeti Special Projects DH','Fox 40 Factory GRIP X2','Fox DHX2 coil',
     'Shimano Saint 220mm','SRAM X01 DH','DT Swiss EX 1700 Spline MX','Maxxis Assegai (F) + Minion DHR II (R)',
     'Fox Racing','Fox Racing / Oakley','Fox Racing / Burgtec','','Pinkbike / Yeti Fox'),

    ('M','Richard','RUDE JR','🇺🇸','Yeti / Fox Factory Race Team',
     'Yeti Special Projects DH','Fox 40 Factory GRIP X2','Fox DHX2 coil',
     'Shimano Saint 220mm','SRAM X01 DH','DT Swiss EX 1700 Spline MX','Maxxis Assegai (F) + Minion DHR II (R)',
     'Fox Racing','Fox Racing / Oakley','Fox Racing / Burgtec','','Pinkbike / Yeti Fox'),

    ('M','Bode','BURKE','🇺🇸','Crestline Speed Shop',
     'Crestline RS 205 VHP (carbon 205mm)','RockShox Boxxer Ultimate','RockShox Super Deluxe Coil Ultimate',
     'SRAM Maven','SRAM X01 DH','Reserve carbon DH','Continental Kryptotal (F) + Trail King (R)',
     'Bell','Bell / Oakley','OneUp / Renthal','Partner: Reserve Wheels, Continental, SRAM/RockShox','Pinkbike / Crestline'),

    ('M','Martin','MAES','🇧🇪','Orbea FMD Racing',
     'Orbea Rallon DH','Fox 40 GRIP X2 200mm','Fox Float X2 Factory 200mm',
     'Shimano XTR 200mm','Shimano Saint 1x7','OQUO MC32 Team MX','Maxxis Assegai 29 (F) + Minion DHR II 27.5 (R)',
     'Poc','Poc / Oakley','Ritchey / Orbea','Mullet 29/27.5','Pinkbike / Orbea FMD'),

    ('M','Stefano','INTROZZI','🇮🇹','Privateer',
     'À confirmer','À confirmer','À confirmer',
     'À confirmer','À confirmer','À confirmer','À confirmer',
     'À confirmer','À confirmer','À confirmer','Privateer IT — pas sur liste UCI 2026',''),

    ('M','Antoine','ROGGE','🇧🇪','Privateer',
     'À confirmer','À confirmer','À confirmer',
     'À confirmer','À confirmer','À confirmer','À confirmer',
     'À confirmer','À confirmer','À confirmer','Privateer BE — pas sur liste UCI 2026',''),

    ('M','Angel','SUAREZ ALONSO','🇪🇸','Privateer',
     'À confirmer','À confirmer','À confirmer',
     'À confirmer','À confirmer','À confirmer','À confirmer',
     'À confirmer','À confirmer','À confirmer','Privateer ES — pas sur liste UCI 2026',''),

    ('M','George','MADLEY','🇬🇧','Privateer',
     'À confirmer','À confirmer','À confirmer',
     'À confirmer','À confirmer','À confirmer','À confirmer',
     'À confirmer','À confirmer','À confirmer','Privateer GB — pas sur liste UCI 2026',''),
]

WOMEN = [
    ('F','Marine','CABIROU','🇫🇷','Canyon DH Racing',
     'Canyon Sender CFR (high-pivot carbon)','RockShox Boxxer Ultimate','RockShox Vivid Coil Ultimate',
     'SRAM Maven Silver','SRAM X01 DH','DT Swiss EX 1700 Spline MX','Maxxis Assegai (F) + Minion DHR II (R)',
     'Troy Lee Designs','Troy Lee Designs / Oakley','Race Face','Mullet option','Pinkbike / Canyon'),

    ('F','Valentina','HÖLL','🇦🇹','Commencal Schwalbe by Les Orres',
     'Commencal Supreme DH V5','Fox 40 Factory GRIP X2','Fox Float X2 Factory',
     'SRAM Maven Silver','SRAM X01 DH','Crankbrothers Synthesis DH','Schwalbe Magic Mary (F) + Hans Dampf (R)',
     'Leatt','Leatt / Oakley','Race Face','Les Orres team = Schwalbe tires','Pinkbike / Commencal'),

    ('F','Lisa','BAUMANN','🇩🇪','Commencal Schwalbe by Les Orres',
     'Commencal Supreme DH V5','Fox 40 Factory GRIP X2','Fox Float X2 Factory',
     'SRAM Maven Silver','SRAM X01 DH','Crankbrothers Synthesis DH','Schwalbe Magic Mary (F) + Hans Dampf (R)',
     'Leatt','Leatt / Oakley','Race Face','','Pinkbike / Commencal'),

    ('F','Myriam','NICOLE','🇫🇷','Commencal / Muc-Off by Riding Addiction',
     'Commencal Supreme DH V5','Fox 40 Factory GRIP X2','Fox Float X2 Factory',
     'SRAM Maven','SRAM X01 DH','Crankbrothers Synthesis DH','Schwalbe Magic Mary (F) + Tacky Chan (R)',
     'Leatt','Leatt / Oakley','Race Face','','Pinkbike / Commencal'),

    ('F','Nina','HOFFMANN','🇩🇪','Santa Cruz Syndicate',
     'Santa Cruz V10','Fox 40 Factory GRIP X2','Fox Float X2 Factory',
     'Shimano Saint','SRAM X01 DH','Reserve 30|HD carbon','Maxxis Assegai (F) + Minion DHR II (R)',
     'Bell','Bell / Oakley','Race Face','','Pinkbike / Santa Cruz'),

    ('F','Eliana','HULSEBOSCH','🇳🇱','Santa Cruz Syndicate',
     'Santa Cruz V10','Fox 40 Factory GRIP X2','Fox Float X2 Factory',
     'Shimano Saint','SRAM X01 DH','Reserve 30|HD carbon','Maxxis Assegai (F) + Minion DHR II (R)',
     'Bell','Bell / Oakley','Race Face','','Pinkbike / Santa Cruz'),

    ('F','Jess','BLEWITT','🇳🇿','Scott Downhill Factory',
     'Scott Gambler RC (6-bar carbon, 210mm)','Fox 40 Factory GRIP X2','Fox DHX2 Factory',
     'SRAM Maven Silver','SRAM X01 DH','DT Swiss EX 1700 Spline MX','Michelin DH22 (F) + Wild Enduro (R)',
     'Scott','Scott / Oakley','Syncros','','Pinkbike / Scott'),

    ('F','Sacha','EARNEST','🇺🇸','Trek - Unbroken DH',
     'Trek Session DH','RockShox Boxxer Factory','RockShox Vivid Coil Factory',
     'SRAM Maven Silver','SRAM X01 DH','Bontrager Line DH','Maxxis Assegai (F) + Minion DHR II (R)',
     'Fox Racing','Fox Racing / Oakley','Burgtec','','Pinkbike / Trek'),

    ('F','Harriet','HARNDEN','🇬🇧','AON Racing',
     'Sego prototype','Manitou Dorado Expert 200mm','Mara coil',
     'Hayes Dominion A4','SRAM X01 DH','Reynolds Blackline DH carbon','Maxxis Assegai (F) + Minion DHR II (R)',
     'Bell','Bell / Oakley','Renthal 780mm','Sego = marque AON Racing propriétaire','Pinkbike / AON Racing'),

    ('F','Phoebe','GALE','🇬🇧','Orbea FMD Racing',
     'Orbea Rallon DH','Fox 40 GRIP X2 200mm','Fox Float X2 Factory 200mm',
     'Shimano XTR 200mm','Shimano Saint 1x7','OQUO MC32 Team MX','Maxxis Assegai 29 (F) + Minion DHR II 27.5 (R)',
     'Poc','Poc / Oakley','Ritchey / Orbea','Mullet 29/27.5','Pinkbike / Orbea FMD'),

    ('F','Tahnée','SEAGRAVE','🇬🇧','Orbea FMD Racing',
     'Orbea Rallon DH','Fox 40 GRIP X2 200mm','Fox Float X2 Factory 200mm',
     'Shimano XTR 200mm','Shimano Saint 1x7','OQUO MC32 Team MX','Maxxis Assegai 29 (F) + Minion DHR II 27.5 (R)',
     'Poc','Poc / Oakley','Ritchey / Orbea','Mullet 29/27.5','Pinkbike / Orbea FMD'),

    ('F','Jenna','HASTINGS','🇨🇦','Pivot Factory Racing',
     'Pivot Phoenix FF','Fox 40 Factory GRIP X2','Fox DHX2 Factory coil',
     'SRAM Maven Silver','SRAM X01 DH','Burgtec EVO-D MK4','Maxxis Assegai (F) + Minion DHR II (R)',
     'Fox Racing','Fox Racing / Oakley','Race Face','','Pinkbike / Pivot'),

    ('F','Matilda','MELTON','🇺🇸','Yeti / Fox Factory Race Team',
     'Yeti Special Projects DH','Fox 40 Factory GRIP X2','Fox DHX2 coil',
     'Shimano Saint 220mm','SRAM X01 DH','DT Swiss EX 1700 Spline MX','Maxxis Assegai (F) + Minion DHR II (R)',
     'Fox Racing','Fox Racing / Oakley','Fox Racing / Burgtec','','Pinkbike / Yeti Fox'),

    ('F','Anna','NEWKIRK','🇺🇸','Frameworks Racing / TRP',
     'Frameworks carbon DH','Fox 40 Factory GRIP X2','Fox Factory coil',
     'TRP DH-EVO 4-piston 220mm','TRP / 5DEV cranks','ENVE carbon DH','Continental DHF (F) + DHR II (R)',
     'Troy Lee Designs','Troy Lee Designs / Oakley','Burgtec','','Pinkbike / Frameworks'),

    ('F','Gloria','SCARSI','🇮🇹','MS-Racing',
     'Zerode G3 (Pinion gearbox + belt)','SR Suntour Rux (custom damper)','Mara Piggy Back coil',
     'Hope Tech 4 (TR4 F + GR4 R) 203mm','Pinion C1.12 + Gates CDX belt','Reynolds Blackline DH carbon','Maxxis Assegai (F) + Minion DHR II (R)',
     'Bell','Bell / Oakley','Renthal 780mm','Gearbox + belt drive, pas de dérailleur','Pinkbike / MS-Racing'),

    ('F','Valentina','ROA SANCHEZ','🇨🇴','MS-Racing',
     'Zerode G3 (Pinion gearbox + belt)','SR Suntour Rux (custom damper)','Mara Piggy Back coil',
     'Hope Tech 4 (TR4 F + GR4 R) 203mm','Pinion C1.12 + Gates CDX belt','Reynolds Blackline DH carbon','Maxxis Assegai (F) + Minion DHR II (R)',
     'Bell','Bell / Oakley','Renthal 780mm','','Pinkbike / MS-Racing'),

    ('F','Veronika','WIDMANN','🇮🇹','Unno Factory Racing DH',
     'Unno Ever (proto alum. lié, adjustable)','SR Suntour RUX (custom damper)','SR Suntour coil (prototype)',
     'TRP DH-EVO 4-piston 220mm','TRP drivetrain','Reserve DH carbon + DT Swiss 350 hubs','Vittoria Mazza (prototype casing)',
     'Poc','Poc / Oakley','Renthal 770mm / 50mm stem','Prototype: géométrie & progression ajustables','Pinkbike bike check Chapelet 2026'),

    ('F','Lisa','BOULADOU','🇫🇷','Santa Cruz Burgtec by Goodman',
     'Santa Cruz V10','Fox 40 Factory GRIP X2','Fox Float X2 Factory',
     'SRAM Maven','SRAM X01 DH','Burgtec EVO-D MK4','Maxxis Assegai (F) + Minion DHR II (R)',
     'Leatt','Leatt / Oakley','Race Face / Burgtec','','Pinkbike / Santa Cruz Burgtec'),

    ('F','Gracey','HEMSTREET','🇨🇦','Norco x Adidas Race Division',
     'Norco Aurum HSP prototype','Fox 40 RAD 203mm','Fox DHX2 coil',
     'Shimano XTR 4-piston 220mm','Shimano XTR 12sp','Crankbrothers Synthesis Carbon DH (Chris King hubs)','Maxxis Assegai (F) + Minion DHF (R)',
     'Poc','Poc / Adidas','OneUp Components','','Pinkbike / Norco'),

    ('F','Frida Helena','RØNNING','🇳🇴','Crestline Speed Shop',
     'Crestline RS 205 VHP (carbon 205mm)','RockShox Boxxer Ultimate','RockShox Super Deluxe Coil Ultimate',
     'SRAM Maven','SRAM X01 DH','Reserve carbon DH','Continental Kryptotal (F) + Trail King (R)',
     'Fox Racing','Fox Racing / Oakley','OneUp / Renthal','Partner: Reserve Wheels, Continental, SRAM/RockShox','Pinkbike / Crestline'),

    ('F','Laïs','BONNAURE','🇫🇷','Privateer / VTT Luberon',
     'À confirmer','À confirmer','À confirmer',
     'À confirmer','À confirmer','À confirmer','À confirmer',
     'À confirmer','À confirmer','À confirmer','Privateer FR — équipe locale VTT Luberon',''),
]

# ── BUILD WORKBOOK ────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = '🏁 Équipements 2026'

def fill(color): return PatternFill('solid', start_color=color, end_color=color)
def font(color=C_WHITE, bold=False, sz=10):
    return Font(name='Arial', color=color, bold=bold, size=sz)
def align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

thin = Side(style='thin', color='DDDDDD')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── TITLE ROW ──
ws.row_dimensions[1].height = 36
ws.merge_cells('A1:Q1')
c = ws['A1']
c.value = '🏁 UCI MTB DH 2026 — Équipements Riders'
c.font = Font(name='Arial', color=C_LIME, bold=True, size=14)
c.fill = fill(C_BG)
c.alignment = align()

# ── HEADER ROW ──
ws.row_dimensions[2].height = 30
for ci, col in enumerate(COLS, 1):
    c = ws.cell(row=2, column=ci, value=col)
    c.font = font(C_BG, bold=True, sz=9)
    c.fill = fill(C_LIME)
    c.alignment = align(wrap=True)
    c.border = border

# ── SEPARATOR HELPER ──
def write_separator(row, label, color):
    ws.row_dimensions[row].height = 20
    ws.merge_cells(f'A{row}:Q{row}')
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(name='Arial', color=C_LIME, bold=True, size=10)
    c.fill = fill(color)
    c.alignment = align()

# ── DATA HELPER ──
def write_rider(row, data, bg):
    ws.row_dimensions[row].height = 22
    for ci, val in enumerate(data, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.fill = fill(bg)
        c.font = font('333333', sz=9)
        c.alignment = align('left', wrap=True)
        c.border = border
    # center Genre & Flag
    ws.cell(row=row, column=1).alignment = align('center')
    ws.cell(row=row, column=4).alignment = align('center')

# ── WRITE MEN ──
write_separator(3, '👨 HOMMES', C_SEP_M)
r = 4
for i, rider in enumerate(MEN):
    bg = C_MEN_A if i % 2 == 0 else C_MEN_B
    write_rider(r, rider, bg)
    r += 1

# ── WRITE WOMEN ──
write_separator(r, '👩 FEMMES', C_SEP_W)
r += 1
for i, rider in enumerate(WOMEN):
    bg = C_WOM_A if i % 2 == 0 else C_WOM_B
    write_rider(r, rider, bg)
    r += 1

# ── COLUMN WIDTHS ──
for ci, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

# ── FREEZE PANES ──
ws.freeze_panes = 'A3'

# ── SAVE ──
out = '/sessions/dreamy-vibrant-volta/mnt/freeride/Equipements_Riders_2026_FULL.xlsx'
wb.save(out)
print(f'Saved: {out}  ({r-1} data rows)')
