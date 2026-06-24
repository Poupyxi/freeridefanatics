#!/usr/bin/env python3
"""
Freeride Fanatics — Instagram Rider Card Generator
====================================================
Structure attendue (tout dans le même dossier) :
  generate_cards.py
  UCI_DH_2026_Tracker_v3.xlsx
  AntonSC-Regular.ttf
  BebasNeue-Regular.ttf
  background.png          ← fond Figma exporté (1080×1350)
  PPRIDERS/
    @annanewkirk.jpg
    @jacksongoldstone.jpg
    ...
  output/                 ← créé automatiquement

UTILISATION :
  python3 generate_cards.py                    # tous les riders
  python3 generate_cards.py --rider NEWKIRK    # un seul rider
  python3 generate_cards.py --women            # Women Elite
  python3 generate_cards.py --men              # Men Elite
"""

import argparse
import csv
import io
import ssl
import urllib.request
from pathlib import Path
from PIL import Image, ImageDraw, ImageFont, ImageFilter
import openpyxl

# ── CONFIGURATION ─────────────────────────────────────────────────────────────

BASE_DIR   = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "output"
LOGOS_DIR  = BASE_DIR / "logos"

PHOTOS_DIR = next((d for d in [BASE_DIR / "PPRIDERS", BASE_DIR / "photos"] if d.exists()),
                  BASE_DIR / "PPRIDERS")

_excels    = sorted(BASE_DIR.glob("UCI_DH_2026_Tracker*.xlsx"),
                    key=lambda p: p.stat().st_mtime, reverse=True)
EXCEL_FILE = _excels[0] if _excels else BASE_DIR / "UCI_DH_2026_Tracker_v3.xlsx"

# ── GOOGLE SHEETS ─────────────────────────────────────────────────────────────
# Colle ici l'ID de ton Google Sheet (la partie entre /d/ et /edit dans l'URL)
GSHEET_ID = ""

BACKGROUND = next((p for p in [
    BASE_DIR / "background.png",
    BASE_DIR / "RIDERS.png",
    BASE_DIR / "assets" / "background.png",
] if p.exists()), None)

def _find_font(name):
    for p in [BASE_DIR / name, BASE_DIR / "fonts" / name]:
        if p.exists():
            return p
    return BASE_DIR / name

FONT_LABEL_PATH = _find_font("AntonSC-Regular.ttf")
FONT_VALUE_PATH = _find_font("BebasNeue-Regular.ttf")

# ── TRADUCTION NATIONALITÉS FR → EN ─────────────────────────────────────────
NATIONALITY_FR_TO_EN = {
    "afghane": "Afghanistan", "albanaise": "Albania", "algérienne": "Algeria",
    "allemande": "Germany", "américaine": "USA", "andorrane": "Andorra",
    "angolaise": "Angola", "argentine": "Argentina", "arménienne": "Armenia",
    "australienne": "Australia", "autrichienne": "Austria", "azerbaïdjanaise": "Azerbaijan",
    "belge": "Belgium", "brésilienne": "Brazil", "britannique": "Great Britain",
    "bulgare": "Bulgaria", "cambodgienne": "Cambodia", "camerounaise": "Cameroon",
    "canadienne": "Canada", "chilienne": "Chile", "chinoise": "China",
    "colombienne": "Colombia", "congolaise": "Congo", "coréenne": "South Korea",
    "costaricienne": "Costa Rica", "croate": "Croatia", "cubaine": "Cuba",
    "danoise": "Denmark", "dominicaine": "Dominican Republic", "écossaise": "Scotland",
    "équatorienne": "Ecuador", "espagnole": "Spain", "estonienne": "Estonia",
    "américaine": "USA", "états-unienne": "USA", "éthiopienne": "Ethiopia",
    "finlandaise": "Finland", "française": "France", "géorgienne": "Georgia",
    "grecque": "Greece", "guatémaltèque": "Guatemala", "hongroise": "Hungary",
    "indienne": "India", "indonésienne": "Indonesia", "irlandaise": "Ireland",
    "islandaise": "Iceland", "israélienne": "Israel", "italienne": "Italy",
    "ivoirienne": "Ivory Coast", "jamaïcaine": "Jamaica", "japonaise": "Japan",
    "jordanienne": "Jordan", "kazakhstanaise": "Kazakhstan", "kényane": "Kenya",
    "lettone": "Latvia", "libanaise": "Lebanon", "lituanienne": "Lithuania",
    "luxembourgeoise": "Luxembourg", "macédonienne": "North Macedonia",
    "malaisienne": "Malaysia", "marocaine": "Morocco", "mexicaine": "Mexico",
    "moldave": "Moldova", "monégasque": "Monaco", "mongole": "Mongolia",
    "monténégrine": "Montenegro", "néerlandaise": "Netherlands", "néo-zélandaise": "New Zealand",
    "nicaraguayenne": "Nicaragua", "nigériane": "Nigeria", "norvégienne": "Norway",
    "ougandaise": "Uganda", "ouzbèke": "Uzbekistan", "pakistanaise": "Pakistan",
    "panaméenne": "Panama", "paraguayenne": "Paraguay", "péruvienne": "Peru",
    "philippine": "Philippines", "polonaise": "Poland", "portugaise": "Portugal",
    "qatarienne": "Qatar", "roumaine": "Romania", "russe": "Russia",
    "rwandaise": "Rwanda", "salvadorienne": "El Salvador", "sénégalaise": "Senegal",
    "serbe": "Serbia", "singapourienne": "Singapore", "slovaque": "Slovakia",
    "slovène": "Slovenia", "somalienne": "Somalia", "sud-africaine": "South Africa",
    "suédoise": "Sweden", "suisse": "Switzerland", "syrienne": "Syria",
    "tanzanienne": "Tanzania", "tchèque": "Czech Republic", "thaïlandaise": "Thailand",
    "tunisienne": "Tunisia", "turque": "Turkey", "ukrainienne": "Ukraine",
    "uruguayenne": "Uruguay", "vénézuélienne": "Venezuela", "vietnamienne": "Vietnam",
    "galloise": "Wales", "anglaise": "England", "zimbabwéenne": "Zimbabwe",
    # Formes masculines/féminines
    "français": "France", "française": "France",
    "américain": "USA", "américaine": "USA",
    "canadien": "Canada", "canadienne": "Canada",
    "australien": "Australia", "australienne": "Australia",
    "allemand": "Germany", "allemande": "Germany",
    "espagnol": "Spain", "espagnole": "Spain",
    "italien": "Italy", "italienne": "Italy",
    "suissesse": "Switzerland",
    "néo-zélandais": "New Zealand", "néo-zélandaise": "New Zealand",
    "belge": "Belgium",
    "norvégien": "Norway", "norvégienne": "Norway",
    "suédois": "Sweden", "suédoise": "Sweden",
    "néerlandais": "Netherlands", "néerlandaise": "Netherlands",
    "autrichien": "Austria", "autrichienne": "Austria",
    "japonais": "Japan", "japonaise": "Japan",
    "chilien": "Chile", "chilienne": "Chile",
    "colombien": "Colombia", "colombienne": "Colombia",
    "brésilien": "Brazil", "brésilienne": "Brazil",
    "mexicain": "Mexico", "mexicaine": "Mexico",
    "polonais": "Poland", "polonaise": "Poland",
    "tchèque": "Czech Republic",
    "slovaque": "Slovakia", "slovaque": "Slovakia",
    "hongrois": "Hungary", "hongroise": "Hungary",
    "croate": "Croatia",
    "serbe": "Serbia",
    "russe": "Russia",
    "ukrainien": "Ukraine", "ukrainienne": "Ukraine",
    "portugais": "Portugal", "portugaise": "Portugal",
    "danois": "Denmark", "danoise": "Denmark",
    "finlandais": "Finland", "finlandaise": "Finland",
    "irlandais": "Ireland", "irlandaise": "Ireland",
    "écossais": "Scotland", "écossaise": "Scotland",
    "gallois": "Wales", "galloise": "Wales",
    "anglais": "England", "anglaise": "England",
    "sud-africain": "South Africa", "sud-africaine": "South Africa",
    "argentin": "Argentina", "argentine": "Argentina",
    "péruvien": "Peru", "péruvienne": "Peru",
    # Noms directs (cas où le tableur a déjà le nom du pays)
    "france": "France", "allemagne": "Germany", "suisse": "Switzerland",
    "espagne": "Spain", "italie": "Italy", "belgique": "Belgium",
    "pays-bas": "Netherlands", "états-unis": "USA", "royaume-uni": "Great Britain",
    "autriche": "Austria", "canada": "Canada", "australie": "Australia",
    "nouvelle-zélande": "New Zealand", "japon": "Japan", "brésil": "Brazil",
    "afrique du sud": "South Africa", "chili": "Chile", "colombie": "Colombia",
    "pérou": "Peru", "mexique": "Mexico", "norvège": "Norway",
    "suède": "Sweden", "danemark": "Denmark", "finlande": "Finland",
    "irlande": "Ireland", "portugal": "Portugal", "pologne": "Poland",
    "russie": "Russia", "ukraine": "Ukraine", "croatie": "Croatia",
    "slovénie": "Slovenia", "tchéquie": "Czech Republic", "hongrie": "Hungary",
    "slovaquie": "Slovakia", "roumanie": "Romania", "bulgarie": "Bulgaria",
    "serbie": "Serbia", "grèce": "Greece", "turquie": "Turkey",
    "israël": "Israel", "maroc": "Morocco", "tunisie": "Tunisia",
    "chine": "China", "corée du sud": "South Korea", "inde": "India",
    "indonésie": "Indonesia", "thaïlande": "Thailand", "vietnam": "Vietnam",
    "philippines": "Philippines", "malaisie": "Malaysia",
    "grande-bretagne": "Great Britain", "angleterre": "England",
    "écosse": "Scotland", "pays de galles": "Wales",
}

def translate_nationality(val):
    """Traduit une nationalité française en anglais. Laisse inchangé si déjà en anglais."""
    if not val:
        return val
    key = val.strip().lower()
    return NATIONALITY_FR_TO_EN.get(key, val)

# Traductions mot-à-mot pour les champs libres (palmares, ville)
# IMPORTANT : les phrases longues doivent être AVANT les mots courts pour éviter les doubles traductions
_FR_WORDS = [
    # === Phrases spécifiques (en premier) ===
    ("Championnats du Monde",    "World Championships"),
    ("Championne du Monde",      "World Champion"),
    ("Champion du Monde",        "World Champion"),
    ("Championne d'Europe",      "European Champion"),
    ("Champion d'Europe",        "European Champion"),
    ("Championne d'Australie",   "Australian Champion"),
    ("Champion d'Australie",     "Australian Champion"),
    ("Championne de Suisse",     "Swiss Champion"),
    ("Champion de Suisse",       "Swiss Champion"),
    ("Championne d'Italie",      "Italian Champion"),
    ("Champion d'Italie",        "Italian Champion"),
    ("Championne de France",     "French Champion"),
    ("Champion de France",       "French Champion"),
    ("Championne d'Allemagne",   "German Champion"),
    ("Champion d'Allemagne",     "German Champion"),
    ("Championne des USA",           "USA Champion"),
    ("Champion des USA",             "USA Champion"),
    ("Championne de Nouvelle-Zélande", "New Zealand Champion"),
    ("Champion de Nouvelle-Zélande",   "New Zealand Champion"),
    ("Championne de New Zealand",    "New Zealand Champion"),
    ("Champion de New Zealand",      "New Zealand Champion"),
    ("Championne d'Angleterre",      "English Champion"),
    ("Champion d'Angleterre",        "English Champion"),
    ("Coupe du Monde",           "World Cup"),
    ("Coupe du monde",           "World Cup"),
    ("1ère victoire",            "1st win"),
    ("1ère femme",               "1st woman"),
    ("1ère saison",              "1st season"),
    ("Victoires WC",             "WC wins"),
    ("Victoire WC",              "WC win"),
    ("Vainqueur WC",             "WC winner"),
    ("Vainqueur",                "Winner"),
    ("vainqueur",                "winner"),
    ("Victoire",                 "Win"),
    ("victoire",                 "win"),
    ("Multiple podiums",         "Multiple podiums"),
    ("multiple podiums",         "multiple podiums"),
    ("Multiple tops",            "Multiple top"),
    ("Jumeau de",                "Twin brother of"),
    ("Frère d'",                 "Brother of "),
    ("Frère de",                 "Brother of"),
    ("Pilier",                   "Key rider"),
    ("années Coupe",             "years on WC"),
    ("saison élite",             "elite season"),
    ("saison",                   "season"),
    ("années",                   "years"),
    ("complète",                 "full"),
    ("À confirmer",              "TBC"),
    ("à confirmer",              "TBC"),
    ("1er",                      "1st"),
    ("2ème",                     "2nd"),
    ("3ème",                     "3rd"),
    ("4ème",                     "4th"),
    ("5ème",                     "5th"),
    ("élite",                    "elite"),
    # === Pays / régions (après les phrases pour éviter doubles traductions) ===
    ("Nouvelle-Galles-du-Sud",   "New South Wales"),
    ("Colombie-Britannique",     "British Columbia"),
    ("Nouvelle-Zélande",         "New Zealand"),
    ("Forêt de Dean",            "Forest of Dean"),
    ("Adelaïde",                 "Adelaide"),
    ("Royaume-Uni",              "UK"),
    ("Angleterre",               "England"),
    ("Écosse",                   "Scotland"),
    ("Pays de Galles",           "Wales"),
    ("Australie",                "Australia"),
    ("Autriche",                 "Austria"),
    ("Allemagne",                "Germany"),
    ("Espagne",                  "Spain"),
    ("Italie",                   "Italy"),
    ("Suisse",                   "Switzerland"),
    ("Thuringe",                 "Thuringia"),
    ("Trentin-Haut-Adige",       "South Tyrol"),
    ("Alpes-Maritimes",          "Alpes-Maritimes"),
    ("Lombardie",                "Lombardy"),
]

def translate_text(val):
    """Remplace les termes français courants par leur équivalent anglais."""
    if not val:
        return val
    for fr, en in _FR_WORDS:
        val = val.replace(fr, en)
    return val

# ── CORRESPONDANCE MARQUES → LOGOS ──────────────────────────────────────────
# Clé = mot-clé dans la valeur Excel (insensible à la casse)
# Valeur = nom du fichier PNG dans logos/
BRAND_MAP = {
    "canyon":          "canyon.png",
    "santa cruz":      "santacruz.png",
    "santacruz":       "santacruz.png",
    "commencal":       "commencal.png",
    "commençal":       "commencal.png",
    "specialized":     "specialized.png",
    "scott":           "scott.png",
    "trek":            "trek.png",
    "yeti":            "yeti.png",
    "pivot":           "pivot.png",
    "mondraker":       "mondraker.png",
    "atherton":        "atherton.png",
    "giant":           "giant.png",
    "rockshox":        "rockshox.png",
    "rock shox":       "rockshox.png",
    "fox":             "fox.png",
    "marzocchi":       "marzocchi.png",
    "ohlins":          "ohlins.png",
    "öhlins":          "ohlins.png",
    "sram":            "sram.png",
    "shimano":         "shimano.png",
    "trp":             "trp.png",
    "dt swiss":        "dtswiss.png",
    "reserve":         "reserve.png",
    "industry nine":   "industrynine.png",
    "industrynine":    "industrynine.png",
    "roval":           "roval.png",
    "bontrager":       "bontrager.png",
    "maxxis":          "maxxis.png",
    "schwalbe":        "schwalbe.png",
    "michelin":        "michelin.png",
    "continental":     "continental.png",
    "troy lee":        "troyleedesigns.png",
    "troyleedesigns":  "troyleedesigns.png",
    "poc":             "poc.png",
    "100%":            "100percent.png",
    "100percent":      "100percent.png",
    "kali":            "kali.png",
    "oakley":          "oakley.png",
    "cushcore":        "cushcore.png",
    "burgtec":         "burgtec.png",
    "oneup":           "oneup.png",
    "one up":          "oneup.png",
    "rideframeworks":  "frameworks.png",
    "ride frameworks": "frameworks.png",
    "redbull":         "redbull.svg",
    "red bull":        "redbull.svg",
    "saalbach":        "saalbach.svg",
    "ergon":           "ergon.svg",
    "five ten":        "fiveten.svg",
    "fiveten":         "fiveten.svg",
    "adidas five ten": "fiveten.svg",
    "gopro":           "gopro.svg",
    "go pro":          "gopro.svg",
    "cupra":           "cupra.svg",
}

LOGO_H         = 50    # hauteur des logos en pixels (largeur proportionnelle auto)
LOGO_GAP       = 12    # espace entre les logos
LOGO_Y         = 1200  # position verticale (axe Y)
LOGO_X         = None  # position horizontale (axe X) :
                        #   None = centré automatiquement dans le panneau gris
                        #   600  = démarre à x=600 (bord gauche panneau)
LOGO_DIRECTION = "row" # disposition : "row" = côte à côte | "col" = empilés verticalement

# ── SPONSORS MIS EN AVANT ────────────────────────────────────────────────────
# Liste les marques à afficher sur les cartes.
# Seules ces marques apparaîtront EN PLUS le rider doit les utiliser.
# Laisse vide [] pour afficher TOUS les sponsors détectés.
# Exemple : FEATURED_BRANDS = ["fox", "maxxis", "shimano"]
FEATURED_BRANDS = [
    "fox",
    "rockshox",
    "maxxis",
    "shimano",
    "sram",
    "poc",
    "oakley",
    "troy lee",
    "100%",
    "rideframeworks",
]

# ── COULEURS & DIMENSIONS ────────────────────────────────────────────────────

C_LIME   = "#C8D400"
C_WHITE  = "#FFFFFF"
C_PANEL  = "#4A4A4A"
C_BORDER = "#B8CC00"

W, H = 1080, 1350

# Position des textes (panneau droit)
TEXT_X   = 580    # ← ajuste si trop à gauche/droite
TEXT_TOP = 80    # ← ajuste si trop haut/bas
GAP      = 50     # espace entre chaque section

# Tailles police
SZ_LABEL    = 36
SZ_VALUE    = 54
SZ_VALUE_SM = 40
SZ_PALMARES = 32

# Seuil luminosité pour détecter la zone blanche du background (0-255)
WHITE_THRESHOLD = 200

# Position de la photo
# PHOTO_OFFSET_X : déplace la photo horizontalement dans la zone blanche
#   0   = centré
#  +100 = décale vers la droite (montre plus le côté gauche du sujet)
#  -100 = décale vers la gauche (montre plus le côté droit du sujet)
PHOTO_OFFSET_X = -200
PHOTO_OFFSET_Y = 0    # même principe vertical
PHOTO_ZOOM     = 1.0  # zoom photo : 1.0 = normal, 1.5 = 50% plus grand, 0.7 = réduit

# ── POLICES ──────────────────────────────────────────────────────────────────

FLAGS_DIR = BASE_DIR / "flags"

_ssl_ctx = ssl.create_default_context()
_ssl_ctx.check_hostname = False
_ssl_ctx.verify_mode   = ssl.CERT_NONE

def _emoji_to_country_code(flag_emoji):
    """Convertit un emoji drapeau (🇫🇷) en code pays 2 lettres (fr)."""
    try:
        chars = [c for c in flag_emoji if '\U0001F1E6' <= c <= '\U0001F1FF']
        if len(chars) >= 2:
            code = ''.join(chr(ord(c) - 0x1F1E6 + ord('A')) for c in chars[:2])
            return code.lower()
    except Exception:
        pass
    return None

def get_flag_image(flag_emoji, height=32):
    """
    Retourne une Image PIL du drapeau.
    Télécharge depuis flagcdn.com et met en cache dans flags/.
    """
    code = _emoji_to_country_code(flag_emoji)
    if not code:
        return None

    FLAGS_DIR.mkdir(exist_ok=True)
    cache_path = FLAGS_DIR / f"{code}.png"

    if not cache_path.exists():
        url = f"https://flagcdn.com/h{height}/{code}.png"
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=10, context=_ssl_ctx) as r:
                cache_path.write_bytes(r.read())
        except Exception as e:
            print(f"    ⚠️  Drapeau '{code}' non téléchargé : {e}")
            return None

    try:
        img = Image.open(cache_path).convert("RGBA")
        ratio = height / img.height
        nw = max(1, int(img.width * ratio))
        return img.resize((nw, height), Image.LANCZOS)
    except Exception:
        return None


def load_fonts():
    fallbacks = [
        "/Library/Fonts/Arial Bold.ttf",
        "/System/Library/Fonts/Helvetica.ttc",
    ]
    def get(path, size):
        try:
            return ImageFont.truetype(str(path), size)
        except:
            print(f"⚠️  Police non trouvée : {Path(path).name}")
            for fb in fallbacks:
                try: return ImageFont.truetype(fb, size)
                except: pass
            return ImageFont.load_default()
    # Police emoji système (macOS)
    emoji_font = None
    for ep in ["/System/Library/Fonts/Apple Color Emoji.ttc",
               "/System/Library/Fonts/AppleColorEmoji.ttf"]:
        try:
            emoji_font = ImageFont.truetype(ep, SZ_VALUE_SM)
            break
        except:
            pass

    return {
        "label":    get(FONT_LABEL_PATH, SZ_LABEL),
        "value":    get(FONT_VALUE_PATH, SZ_VALUE),
        "value_sm": get(FONT_VALUE_PATH, SZ_VALUE_SM),
        "palmares": get(FONT_VALUE_PATH, SZ_PALMARES),
        "emoji":    emoji_font,
    }

# ── GOOGLE SHEETS READER ─────────────────────────────────────────────────────

def _fetch_gsheet_csv(sheet_name):
    """Télécharge une feuille Google Sheets en CSV. Retourne une liste de lignes (listes)."""
    import urllib.parse
    encoded = urllib.parse.quote(sheet_name)
    url = f""
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10, context=_ssl_ctx) as r:
            content = r.read().decode("utf-8")
        rows = list(csv.reader(io.StringIO(content)))
        if rows and rows[0]:  # vérifie que le contenu est valide
            return rows
    except Exception as e:
        print(f"  ⚠️  Google Sheets '{sheet_name}' inaccessible : {e}")
    return None

def _gsheet_row_to_list(row, length):
    """Complète une ligne CSV jusqu'à la longueur demandée."""
    row = list(row)
    while len(row) < length:
        row.append("")
    return row

def load_profiles_from_gsheet():
    """
    Charge les profils depuis Google Sheets — source unique de vérité.

    Structure 👤 Profils :
      0=G  1=First Name  2=Last Name  3=🏳️  4=Nationality  5=Hometown
      6=Date of Birth  7=Age  8=Instagram  9=Achievements  10=Team  11=Sponsors

    Structure 🏆 Riders :
      0=G  1=First Name  2=Last Name  3=Bike  4=Fork  5=Shock  6=Brakes
      7=Drivetrain  8=Wheels  9=Tires  10=Helmet  11=Goggles  12=Cockpit  13=Inserts
    """
    print(f"☁️  Google Sheets : {GSHEET_ID}")

    riders_rows  = _fetch_gsheet_csv("🏆 Riders")
    profils_rows = _fetch_gsheet_csv("👤 Profils")

    if not profils_rows:
        return None

    # ── Riders → équipement uniquement ───────────────────────────────────────
    rider_equip = {}
    if riders_rows:
        for row in riders_rows:
            row = _gsheet_row_to_list(row, 14)
            if row[0].strip().upper() not in ("F", "M"):
                continue
            nom = row[2].strip()
            if not nom:
                continue
            # cols 3-13 = équipements (Bike → Inserts)
            equip_vals = [row[i].strip() for i in range(3, 14) if row[i].strip() and row[i].strip() != "N/A"]
            rider_equip[nom] = equip_vals

    # ── Profils → tout le reste ───────────────────────────────────────────────
    profiles = []
    for row in profils_rows:
        row = _gsheet_row_to_list(row, 15)
        g = row[0].strip().upper()
        if g not in ("F", "M"):
            continue
        nom = row[2].strip()
        if not nom:
            continue

        # Colonne Sponsors (index 11, séparateur ;)
        sponsors_raw = row[11].strip() if len(row) > 11 else ""
        sponsors = [s.strip() for s in sponsors_raw.split(";") if s.strip()] if sponsors_raw else []

        profiles.append({
            "genre":       g,
            "prenom":      row[1].strip(),
            "nom":         nom,
            "flag":        row[3].strip(),
            "nationalite": translate_nationality(row[4].strip()) or "N/A",
            "ville":       translate_text(row[5].strip()) if row[5].strip() else "N/A",
            "birthday":    row[6].strip() or "N/A",
            "age":         row[7].strip() or "?",
            "instagram":   row[8].strip(),
            "palmares":    translate_text(row[9].strip()) if row[9].strip() else "N/A",
            "team":        row[10].strip() or "TBC",
            "sponsors":    sponsors,
            "equipment":   rider_equip.get(nom, []),
            "category":    "Women Elite" if g == "F" else "Men Elite",
        })

    print(f"✅ {len(profiles)} riders chargés depuis Google Sheets")
    if profiles:
        ex = next((p for p in profiles if p["sponsors"]), None)
        if ex:
            print(f"  ex. {ex['nom']} → sponsors: {ex['sponsors']}")
    return profiles

# ── DONNÉES EXCEL (fallback) ──────────────────────────────────────────────────

def load_profiles():
    print(f"📂 Tracker : {EXCEL_FILE.name}")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    ws_riders = wb["🏆 Riders"]
    # Détecte la colonne Sponsors dynamiquement depuis les headers (row 2)
    headers_row = [str(c.value).strip().lower() if c.value else "" for c in ws_riders[2]]
    col_sponsors = headers_row.index("sponsors") if "sponsors" in headers_row else None

    rider_info = {}
    for row in ws_riders.iter_rows(min_row=3, values_only=True):
        if row[0] in ("F", "M") and row[2]:
            equip_vals = [str(row[i]) for i in range(6, min(16, len(row))) if row[i]]
            sponsors_raw = str(row[col_sponsors]).strip() if col_sponsors and len(row) > col_sponsors and row[col_sponsors] else ""
            sponsors = [s.strip() for s in sponsors_raw.split(";") if s.strip()]
            rider_info[row[2]] = {
                "team":      row[5] or "TBC",
                "instagram": row[16] or "" if len(row) > 16 else "",
                "equipment": equip_vals,
                "sponsors":  sponsors,
            }

    ws_profils = wb["👤 Profils"]
    profiles = []
    for row in ws_profils.iter_rows(min_row=3, values_only=True):
        if row[0] in ("F", "M") and row[2]:
            nom  = row[2]
            info = rider_info.get(nom, {})
            profiles.append({
                "genre":       row[0],
                "prenom":      row[1] or "",
                "nom":         nom,
                "flag":        row[3] or "",
                "nationalite": translate_nationality(row[4] or "") or "N/A",
                "ville":       translate_text(str(row[5])) if row[5] else "TBC",
                "birthday":    str(row[6]) if row[6] else "TBC",
                "age":         str(row[7]) if row[7] else "?",
                "palmares":    translate_text(str(row[8])) if row[8] else "TBC",
                "category":    "Women Elite" if row[0] == "F" else "Men Elite",
                "team":        info.get("team", "TBC"),
                "instagram":   info.get("instagram", ""),
                "equipment":   info.get("equipment", []),
                "sponsors":    info.get("sponsors", []),
            })
    return profiles

# ── RECHERCHE PHOTO ───────────────────────────────────────────────────────────

def _ascii(s):
    """Supprime les accents/diacritiques pour comparaison souple."""
    import unicodedata
    return unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode("ascii").lower()

def find_photo(profile):
    instagram = profile.get("instagram", "").lstrip("@").lower()
    nom       = profile["nom"].lower().replace(" ", "_").replace("-", "_")
    prenom    = profile["prenom"].lower().replace(" ", "_").replace("-", "_")
    nom_ascii = _ascii(nom).replace("_","")
    prenom_ascii = _ascii(prenom)

    candidates = []
    if instagram:
        candidates += [f"@{instagram}", instagram]
    candidates += [f"{nom}_{prenom}", f"{prenom}_{nom}", nom, prenom]

    for directory in [d for d in [PHOTOS_DIR, BASE_DIR] if d.exists()]:
        # 1. Correspondance exacte
        for stem in candidates:
            for ext in [".jpg", ".jpeg", ".png"]:
                p = directory / (stem + ext)
                if p.exists():
                    return p
        # 2. Correspondance partielle (avec normalisation unicode)
        for f in sorted(directory.iterdir()):
            if f.suffix.lower() not in [".jpg", ".jpeg", ".png"]:
                continue
            stem_l     = f.stem.lower().lstrip("@")
            stem_ascii = _ascii(stem_l).replace("_","").replace("-","")
            if instagram and (_ascii(instagram) in stem_ascii or stem_ascii in _ascii(instagram)):
                return f
            if nom_ascii and nom_ascii in stem_ascii:
                return f
            if prenom_ascii and prenom_ascii in stem_ascii and nom_ascii in stem_ascii:
                return f
    return None

# ── FOND DE SECOURS ───────────────────────────────────────────────────────────

def make_fallback_bg():
    import random
    random.seed(42)
    img  = Image.new("RGB", (W, H), "#FFFFFF")
    draw = ImageDraw.Draw(img)
    pts  = [(415 + random.randint(-18,18), y) for y in range(0, H+10, 8)]
    draw.polygon([(415,0)] + pts + [(415,H),(W,H),(W,0)], fill=C_PANEL)
    b = 8
    draw.rectangle([b//2, b//2, W-b//2, H-b//2], outline=C_BORDER, width=b)
    return img

# ── GÉNÉRATION CARTE ──────────────────────────────────────────────────────────

def place_photo_on_canvas(img, canvas_w, canvas_h):
    """
    Place la photo sur un canvas blanc.
    PHOTO_OFFSET_X / PHOTO_OFFSET_Y décalent librement la photo.
    Valeurs typiques : -400 à +400
    """
    # Redimensionne pour remplir le canvas + zoom optionnel
    ratio = max(canvas_w / img.width, canvas_h / img.height) * PHOTO_ZOOM
    nw = int(img.width  * ratio)
    nh = int(img.height * ratio)
    img = img.resize((nw, nh), Image.LANCZOS)

    # Position de base : centrée
    x = (canvas_w - nw) // 2 + PHOTO_OFFSET_X
    y = (canvas_h - nh) // 2 + PHOTO_OFFSET_Y

    canvas = Image.new("RGB", (canvas_w, canvas_h), "#FFFFFF")
    canvas.paste(img.convert("RGB"), (x, y))
    return canvas


def composite_photo_into_bg(photo_img, bg_rgb):
    """
    Place la photo DERRIÈRE le background.
    Le background s'affiche par-dessus pour garder le brush stroke et la bordure.
    La photo n'apparaît que dans la zone blanche (pixels clairs du background).
    """
    # Photo déjà positionnée sur canvas par place_photo_on_canvas
    photo_full = photo_img.convert("RGB")

    # Masque : pixels blancs/clairs du background → zone photo
    bg_gray = bg_rgb.convert("L")
    mask = bg_gray.point(lambda x: 255 if x > WHITE_THRESHOLD else 0)
    # Légère atténuation du bord pour un rendu plus propre
    mask = mask.filter(ImageFilter.GaussianBlur(3))

    # Composite : photo là où le background est clair, background ailleurs
    result = Image.composite(photo_full, bg_rgb, mask)
    return result


def extract_brands(equipment_list):
    """
    Détection automatique depuis le tableur Excel.
    Si FEATURED_BRANDS est défini, filtre sur ces marques uniquement.
    """
    seen_files = []
    seen_set   = set()
    featured   = [b.lower() for b in FEATURED_BRANDS]

    for val in equipment_list:
        val_lower = val.lower()
        for keyword, filename in BRAND_MAP.items():
            if keyword not in val_lower:
                continue
            if filename in seen_set:
                continue
            if featured and not any(f in keyword or keyword in f for f in featured):
                continue
            logo_path = LOGOS_DIR / filename
            if logo_path.exists():
                seen_files.append(logo_path)
                seen_set.add(filename)
            break
    return seen_files


def _find_logo_file(stem):
    """Cherche le fichier logo pour un stem donné — SVG prioritaire sur PNG."""
    for ext in [".svg", ".png"]:
        p = LOGOS_DIR / (stem + ext)
        if p.exists():
            return p
    return None

def brands_from_names(names):
    """
    Résout une liste de clés/noms vers les fichiers logo correspondants.
    Stratégie :
      1. Essai direct par stem (SVG > PNG)
      2. Recherche dans BRAND_MAP (keyword matching), puis essai SVG si PNG absent
      3. Recherche partielle dans le dossier logos/
    """
    seen_files = []
    seen_set   = set()

    for name in names:
        name_lower = name.lower()
        found = None

        # 1. Essai direct par stem (ex: "commencal" → commencal.svg ou commencal.png)
        found = _find_logo_file(name_lower)

        # 2. Recherche via BRAND_MAP
        if not found:
            for keyword, filename in BRAND_MAP.items():
                if name_lower in keyword or keyword in name_lower:
                    stem = Path(filename).stem
                    found = _find_logo_file(stem)  # SVG > PNG
                    break

        # 3. Recherche partielle dans logos/ (ex: "fox" trouve "fox-logo.svg")
        if not found and LOGOS_DIR.exists():
            for f in sorted(LOGOS_DIR.iterdir()):
                if f.suffix.lower() in (".svg", ".png") and name_lower in f.stem.lower():
                    found = f
                    break

        if found and found.name not in seen_set:
            seen_files.append(found)
            seen_set.add(found.name)
        elif not found:
            print(f"    ⚠️  Logo introuvable pour '{name}'")

    return seen_files


def draw_logos(card, logo_paths):
    """Colle une rangée de logos en bas du panneau gris."""
    if not logo_paths or not LOGOS_DIR.exists():
        return card

    logos = []
    for p in logo_paths:
        try:
            img = Image.open(p).convert("RGBA")
            ratio = LOGO_H / img.height
            nw = int(img.width * ratio)
            img = img.resize((nw, LOGO_H), Image.LANCZOS)
            logos.append(img)
        except Exception as e:
            print(f"    ⚠️  Logo erreur ({p.name}) : {e}")

    if not logos:
        return card

    panel_left  = TEXT_X - 20
    panel_right = W - 20
    max_width   = panel_right - panel_left

    card_rgba = card.convert("RGBA")

    if LOGO_DIRECTION == "col":
        # ── MODE COLONNE : logos empilés verticalement ──────────────────────
        # Largeur max = largeur du panneau → on redimensionne chaque logo si besoin
        x_start = LOGO_X if LOGO_X is not None else panel_left
        y = LOGO_Y
        for logo in logos:
            # Limite la largeur au panneau
            if logo.width > max_width:
                scale = max_width / logo.width
                logo = logo.resize((max_width, max(1, int(logo.height * scale))), Image.LANCZOS)
            # Centre horizontalement dans le panneau si LOGO_X non forcé
            x = x_start if LOGO_X is not None else panel_left + (max_width - logo.width) // 2
            card_rgba.paste(logo, (x, y), logo)
            y += logo.height + LOGO_GAP

    else:
        # ── MODE LIGNE : logos côte à côte (défaut) ─────────────────────────
        total_w = sum(l.width for l in logos) + LOGO_GAP * (len(logos) - 1)

        # Si trop large : réduit tous proportionnellement
        if total_w > max_width:
            scale = max_width / total_w
            logos = [l.resize((max(1, int(l.width * scale)), max(1, int(l.height * scale))),
                               Image.LANCZOS) for l in logos]
            total_w = sum(l.width for l in logos) + LOGO_GAP * (len(logos) - 1)

        x = LOGO_X if LOGO_X is not None else panel_left + (max_width - total_w) // 2
        y = LOGO_Y
        for logo in logos:
            card_rgba.paste(logo, (x, y), logo)
            x += logo.width + LOGO_GAP

    return card_rgba.convert("RGB")


def draw_wrapped(draw, text, font, color, x, y, max_w):
    words = text.split()
    lines, cur = [], ""
    for w in words:
        test = (cur + " " + w).strip()
        if draw.textbbox((0,0), test, font=font)[2] <= max_w:
            cur = test
        else:
            if cur: lines.append(cur)
            cur = w
    if cur: lines.append(cur)
    for line in lines:
        draw.text((x, y), line, font=font, fill=color)
        y += font.size + 4
    return y


def generate_card(profile, fonts, bg_rgb, forced_sponsors=None):
    # Photo dans la zone blanche du background
    photo_path = find_photo(profile)
    if photo_path:
        try:
            photo        = Image.open(photo_path)
            photo_canvas = place_photo_on_canvas(photo, W, H)
            card         = composite_photo_into_bg(photo_canvas, bg_rgb)
        except Exception as e:
            print(f"    ⚠️  Photo erreur : {e}")
            card = bg_rgb.copy()
    else:
        print(f"    ℹ️  Pas de photo : {profile['prenom']} {profile['nom']}")
        card = bg_rgb.copy()

    # Textes par-dessus
    draw  = ImageDraw.Draw(card)
    max_w = W - TEXT_X - 20

    # Handle Instagram en haut du panneau gris
    handle = profile.get("instagram", "").strip()
    if handle:
        if not handle.startswith("@"):
            handle = "@" + handle
        draw.text((TEXT_X, 20), handle.lower(), font=fonts["label"], fill=C_LIME)

    sections = [
        ("Name",        f"{profile['prenom']} {profile['nom']}",           "value",    None),
        ("Nationality", profile["nationalite"],                              "value_sm", profile.get("flag", "")),
        ("Birthday",    f"{profile['birthday']}  ·  {profile['age']} y.o.", "value_sm", None),
        ("Homestay",    profile["ville"],                                    "value_sm", None),
        ("Category",    profile["category"],                                 "value_sm", None),
        ("Team",        profile["team"],                                     "value_sm", None),
        ("Palmares",    profile["palmares"],                                 "palmares", None),
    ]

    y = TEXT_TOP
    for label, value, fkey, flag in sections:
        draw.text((TEXT_X, y), label.upper(), font=fonts["label"], fill=C_LIME)
        y += fonts["label"].size + 15
        # Drapeau image avant la nationalité
        x_val = TEXT_X
        if flag:
            flag_img = get_flag_image(flag, height=fonts[fkey].size)
            if flag_img:
                # Colle le drapeau aligné verticalement avec le texte
                fy = y + max(0, (fonts[fkey].size - flag_img.height) // 2)
                card.paste(flag_img, (x_val, fy), flag_img)
                x_val += flag_img.width + 8
        y = draw_wrapped(draw, value.upper(), fonts[fkey], C_WHITE, x_val, y, max_w - (x_val - TEXT_X))
        y += GAP

    # Logos sponsors — priorité : 1. forced (app/CLI) 2. colonne Sponsors du sheet 3. auto équipement
    if forced_sponsors is not None:
        logo_paths = brands_from_names(forced_sponsors)
    elif profile.get("sponsors"):
        logo_paths = brands_from_names(profile["sponsors"])
    else:
        logo_paths = extract_brands(profile.get("equipment", []))
    if logo_paths:
        card = draw_logos(card, logo_paths)

    return card

# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--rider",    help="Nom du rider (ex: GOLDSTONE)")
    parser.add_argument("--women",    action="store_true")
    parser.add_argument("--men",      action="store_true")
    parser.add_argument("--sponsors", nargs="*", metavar="MARQUE",
                        help="Logos à forcer (ex: --sponsors fox maxxis rideframeworks). "
                             "Si absent : détection auto depuis le tableur.")
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(exist_ok=True)

    if LOGOS_DIR.exists():
        logo_count = len(list(LOGOS_DIR.glob("*.png")))
        print(f"🏷️  Logos : {logo_count} PNG trouvés dans logos/")
    else:
        print("⚠️  Dossier logos/ absent → aucun logo affiché")

    if BACKGROUND:
        bg = Image.open(BACKGROUND).convert("RGB").resize((W, H), Image.LANCZOS)
        print(f"✅ Fond : {BACKGROUND.name}")
    else:
        print("⚠️  background.png absent → fond généré automatiquement")
        bg = make_fallback_bg()

    fonts    = load_fonts()
    profiles = None
    if GSHEET_ID:
        profiles = load_profiles_from_gsheet()
    if not profiles:
        profiles = load_profiles()
    print(f"✅ {len(profiles)} riders chargés\n")

    if args.rider:
        profiles = [p for p in profiles if args.rider.upper() in p["nom"].upper()]
    if args.women:
        profiles = [p for p in profiles if p["genre"] == "F"]
    if args.men:
        profiles = [p for p in profiles if p["genre"] == "M"]

    print(f"→ {len(profiles)} carte(s) à générer...\n")

    forced = args.sponsors  # None si pas spécifié, [] si --sponsors sans args

    if forced is not None:
        if forced:
            print(f"🏷️  Sponsors forcés : {', '.join(forced)}")
        else:
            print("🏷️  --sponsors vide : aucun logo affiché")

    for p in profiles:
        slug = f"{p['nom'].lower().replace(' ','_')}_{p['prenom'].lower()}"
        out  = OUTPUT_DIR / f"{slug}.jpg"
        print(f"  🎨 {p['prenom']} {p['nom']}")
        card = generate_card(p, fonts, bg, forced_sponsors=forced)
        card.save(out, "JPEG", quality=95)
        print(f"     ✅ {out.name}")

    print(f"\n✅ Terminé — dossier output/")

if __name__ == "__main__":
    main()
